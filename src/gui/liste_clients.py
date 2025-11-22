def afficher_liste_clients(parent):
    """
    Affiche la vue 'Liste des clients' dans `parent`.
    Retourne dict: {"refresh": callable, "search_var": tk.StringVar, "page_size_var": tk.IntVar}
    """
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
    from database.connection import get_connection

    # try to access global session (role)
    try:
        from models.session import session as global_session
    except Exception:
        global_session = None

    # visual config
    CONTENT_BG = "#f6f8fa"
    CARD_BG = "#ffffff"
    CONTOUR_BG = "#e6eef9"
    HEADER_BG = "#eef6ff"
    HEADER_FG = "#0b3d91"
    LABEL_FG = "#1f2937"
    ROW_BG_1 = CARD_BG
    ROW_BG_2 = "#fbfdff"
    ENTRY_BG = "white"
    ENTRY_BORDER = "#ced4da"
    BTN_FG = "white"
    FONT_CELL = ("Segoe UI", 9)
    FONT_LABEL = ("Segoe UI", 10)
    FONT_TITLE = ("Segoe UI", 14, "bold")

    # columns
    COLUMNS = [
        ("id", "ID", 40),
        ("customer_name", "Nom", 160),
        ("customer_TIN", "NIF", 100),
        ("customer_type", "Type", 70),
        ("vat_customer_payer", "TVA", 50),
    ]

    ROW_HEIGHT = 32
    PAGE_SIZES = [10, 15, 20, 50]
    default_page_size = 15


    def _format_cell(col, val):
        if val is None:
            return ""
        if col == "customer_type":
            return "Physique" if str(val) == "1" else "Morale" if str(val) == "2" else str(val)
        if col == "vat_customer_payer":
            return "Oui" if str(val) == "1" else "Non"
        return str(val)

    # session/permission helper
    def _current_user_can_delete():
        try:
            if not global_session:
                return False
            role = getattr(global_session, "role", None)
            if role is None:
                return False
            role = str(role).lower()
            return role in ("admin", "manager")
        except Exception:
            return False

    def _get_full_client_by_id(cid):
        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM client WHERE id = ?", (cid,))
            row = cur.fetchone()
            desc = [d[0] for d in cur.description] if cur.description else []
            conn.close()
            if row:
                return {desc[i]: row[i] for i in range(len(desc))}
        except Exception:
            pass
        return {}

    def _delete_client_by_id(cid, refresh_cb=None, parent_widget=None):
        # Defensive permission check
        if not _current_user_can_delete():
            messagebox.showerror("Permission", "Vous n'êtes pas autorisé à supprimer des clients.", parent=parent_widget)
            return
        if not messagebox.askyesno("Confirmer", "Voulez-vous supprimer ce client ?", parent=parent_widget):
            return
        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("DELETE FROM client WHERE id = ?", (cid,))
            conn.commit()
            conn.close()
            if callable(refresh_cb):
                try: refresh_cb()
                except Exception: pass
            messagebox.showinfo("Supprimé", "Client supprimé ✅", parent=parent_widget)
        except Exception as e:
            messagebox.showerror("Erreur", f"Échec suppression : {e}", parent=parent_widget)

    # clear parent
    for w in parent.winfo_children():
        try: w.destroy()
        except Exception: pass

    parent.grid_columnconfigure(0, weight=1)
    parent.grid_rowconfigure(1, weight=1)
    parent.configure(bg=CONTENT_BG)

    # Header
    header = tk.Frame(parent, bg=CONTENT_BG)
    header.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
    header.grid_columnconfigure(0, weight=1)

    tk.Label(header, text="📋 Liste des clients", font=FONT_TITLE, bg=CONTENT_BG, fg=HEADER_FG).grid(row=0, column=0, sticky="w")

    right = tk.Frame(header, bg=CONTENT_BG)
    right.grid(row=0, column=1, sticky="e")

    # search + page size + export buttons
    search_var = tk.StringVar()
    tk.Label(right, text="Recherche :", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL).pack(side="left", padx=(0,6))
    search_entry = tk.Entry(right, textvariable=search_var, font=FONT_LABEL, width=22,
                             bg=ENTRY_BG, highlightthickness=1, highlightbackground=ENTRY_BORDER, relief="flat")
    search_entry.pack(side="left", padx=(0,8))

    page_size_var = tk.IntVar(value=default_page_size)
    tk.Label(right, text="Taille page:", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL).pack(side="left", padx=(6,4))
    page_size_cb = ttk.Combobox(right, values=PAGE_SIZES, textvariable=page_size_var, width=4, state="readonly")
    page_size_cb.pack(side="left", padx=(0,8))
    page_size_cb.set(default_page_size)

    btn_export_xl = tk.Button(right, text="Export Excel", bg="#16a34a", fg="white", activebackground="#15803d", width=12)
    btn_export_pdf = tk.Button(right, text="Export PDF", bg="#2563eb", fg="white", activebackground="#1e40af", width=12)
    btn_export_xl.pack(side="left", padx=(4,4))
    btn_export_pdf.pack(side="left", padx=(4,0))

    # content card and inner grid
    card = tk.Frame(parent, bg=CARD_BG)
    card.grid(row=1, column=0, sticky="nsew", padx=12, pady=6)
    card.grid_columnconfigure(0, weight=1)
    card.grid_rowconfigure(0, weight=1)

    inner_outer = tk.Frame(card, bg=CONTOUR_BG, bd=1, relief="solid")
    inner_outer.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
    inner_outer.grid_rowconfigure(0, weight=1)
    inner_outer.grid_columnconfigure(0, weight=1)
    inner_outer.grid_columnconfigure(1, weight=0)
    inner_outer.grid_columnconfigure(2, weight=1)

    inner_grid = tk.Frame(inner_outer, bg=CARD_BG)
    inner_outer.grid_columnconfigure(1, weight=1)
    inner_grid.grid(row=0, column=1, sticky="nsew", padx=0, pady=8)
    inner_grid.grid_rowconfigure(0, weight=0)

    # header row
    for ci, (dbcol, label, minw) in enumerate(COLUMNS):
        h = tk.Label(inner_grid, text=label, bg=HEADER_BG, fg=HEADER_FG, font=("Segoe UI", 9, "bold"),
                      anchor="w", padx=6, bd=1, relief="solid", pady=8)
        h.grid(row=0, column=ci, sticky="nsew", padx=0, pady=0)
        if label == "Nom":
            inner_grid.grid_columnconfigure(ci, weight=1, minsize=minw)
        elif label in ("ID", "TVA"):
            inner_grid.grid_columnconfigure(ci, weight=0, minsize=max(30, minw - 5))
        else:
            inner_grid.grid_columnconfigure(ci, weight=0, minsize=minw)

    # actions header column
    act_col = len(COLUMNS)
    ha = tk.Label(inner_grid, text="Actions", bg=HEADER_BG, fg=HEADER_FG, font=("Segoe UI", 9, "bold"),
                  anchor="w", padx=6, bd=1, relief="solid", pady=8)
    ha.grid(row=0, column=act_col, sticky="nsew", padx=0, pady=0)
    inner_grid.grid_columnconfigure(act_col, weight=0, minsize=200)

    # pagination state
    current_page = {"n": 1}
    pagination_info = {"total": 0}

    def _fetch_clients_page(filter_text=None, page=1, page_size=15):
        conn = get_connection()
        cur = conn.cursor()
        cols = [c[0] for c in COLUMNS]
        sql_cols = ", ".join(cols)
        where = ""
        params = []
        if filter_text:
            where = "WHERE customer_name LIKE ? OR customer_TIN LIKE ? OR IFNULL(customer_email,'') LIKE ?"
            q = f"%{filter_text}%"
            params = [q, q, q]
        cur.execute(f"SELECT COUNT(1) FROM client {where}", params)
        total = cur.fetchone()[0]
        offset = (page - 1) * page_size
        cur.execute(f"SELECT {sql_cols} FROM client {where} ORDER BY customer_name LIMIT ? OFFSET ?", params + [page_size, offset])
        rows = cur.fetchall()
        conn.close()
        return rows, total

    def _get_full_client_by_nif(nif):
        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM client WHERE customer_TIN = ?", (nif,))
            row = cur.fetchone()
            desc = [d[0] for d in cur.description] if cur.description else []
            conn.close()
            if row:
                return {desc[i]: row[i] for i in range(len(desc))}
        except Exception:
            pass
        return {}

    def _view_client_modal(client):
        if not client:
            messagebox.showinfo("Détails client", "Aucune donnée")
            return
        lines = [f"{k}: {v}" for k, v in client.items()]
        messagebox.showinfo("Détails client", "\n".join(lines))

    def _edit_client_modal(client, refresh_cb):
        """
        Affiche un modal d'édition local si la fonction externe n'est pas disponible.
        Pré-remplit les champs et propose Enregistrer / Annuler.
        """
        if not client:
            messagebox.showinfo("Édition", "Aucune donnée à éditer")
            return

        # Try centralized editor first
        try:
            from gui.window_client import afficher_formulaire_client
        except Exception:
            afficher_formulaire_client = None

        if afficher_formulaire_client:
            modal = tk.Toplevel(parent)
            modal.transient(parent); modal.grab_set(); modal.title("Éditer le client")
            try:
                afficher_formulaire_client(modal, client)
                modal.wait_window()
                if callable(refresh_cb): refresh_cb()
                return
            except TypeError:
                modal.destroy()

        # Local inline modal editor (fallback)
        dlg = tk.Toplevel(parent)
        dlg.transient(parent)
        dlg.grab_set()
        dlg.title(f"✏️ Éditer client — {client.get('customer_TIN','')}")
        dlg.geometry("800x520")
        dlg.update_idletasks()

        frame = tk.Frame(dlg, bg=CARD_BG, padx=14, pady=14)
        frame.pack(fill="both", expand=True)

        fields = [
            ("Nom du client", "customer_name"),
            ("NIF", "customer_TIN"),
            ("Adresse", "customer_address"),
            ("Téléphone", "customer_phone_number"),
            ("Email", "customer_email"),
        ]

        vars_map = {}
        for i, (label_text, key) in enumerate(fields):
            tk.Label(frame, text=label_text + " :", bg=CARD_BG, fg=LABEL_FG, font=("Segoe UI", 10, "bold")).grid(row=i, column=0, sticky="w", pady=8)
            val = "" if client.get(key) is None else str(client.get(key))
            var = tk.StringVar(value=val)
            ent = tk.Entry(frame, textvariable=var, width=60, font=FONT_LABEL, bg=ENTRY_BG, bd=1, relief="flat", highlightbackground=ENTRY_BORDER, highlightthickness=1)
            ent.grid(row=i, column=1, sticky="w", pady=8, padx=(8,0))
            vars_map[key] = var

        # customer_type and vat checkbox
        row_idx = len(fields)
        tk.Label(frame, text="Type (1=Physique / 2=Morale) :", bg=CARD_BG, fg=LABEL_FG, font=("Segoe UI", 10, "bold")).grid(row=row_idx, column=0, sticky="w", pady=8)
        type_var = tk.StringVar(value=str(client.get("customer_type") or ""))
        type_cb = ttk.Combobox(frame, textvariable=type_var, values=["", "1", "2"], state="normal", width=10, font=FONT_LABEL)
        type_cb.grid(row=row_idx, column=1, sticky="w", pady=8, padx=(8,0))

        row_idx += 1
        vat_val = client.get("vat_customer_payer")
        vat_var = tk.BooleanVar(value=(str(vat_val) == "1" or vat_val is True))
        tk.Checkbutton(frame, text="Assujetti à la TVA", variable=vat_var, bg=CARD_BG).grid(row=row_idx, column=1, sticky="w", pady=4, padx=(8,0))

        # action buttons
        btn_frame = tk.Frame(frame, bg=CARD_BG)
        btn_frame.grid(row=row_idx+1, column=0, columnspan=2, pady=(16,0), sticky="e")

        def _validate_and_save():
            name = vars_map["customer_name"].get().strip()
            nif = vars_map["customer_TIN"].get().strip()
            email = vars_map["customer_email"].get().strip()
            phone = vars_map["customer_phone_number"].get().strip()
            cust_type = type_var.get().strip()
            vat = "1" if vat_var.get() else "0"

            if not name:
                messagebox.showwarning("Validation", "Le nom du client est requis", parent=dlg); return
            if not nif:
                messagebox.showwarning("Validation", "Le NIF est requis", parent=dlg); return

            try:
                conn = get_connection()
                cur = conn.cursor()
                cur.execute("PRAGMA table_info(client)")
                cols_info = cur.fetchall()
                cols = [c["name"] if isinstance(c, sqlite3.Row) else c[1] for c in cols_info]

                update_pairs = []
                params = []
                mapping = {
                    "customer_name": name,
                    "customer_TIN": nif,
                    "customer_address": vars_map["customer_address"].get().strip(),
                    "customer_phone_number": phone,
                    "customer_email": email,
                    "customer_type": cust_type if cust_type else None,
                    "vat_customer_payer": vat
                }
                for col_name, val in mapping.items():
                    if col_name in cols:
                        update_pairs.append(f"{col_name} = ?")
                        params.append(val)
                if not update_pairs:
                    messagebox.showinfo("Mise à jour", "Aucune colonne disponible à mettre à jour dans la table client", parent=dlg)
                    conn.close()
                    return

                original_id = client.get("id")
                if original_id is not None:
                    params.append(original_id)
                    sql = f"UPDATE client SET {', '.join(update_pairs)} WHERE id = ?"
                else:
                    original_nif = client.get("customer_TIN")
                    params.append(original_nif)
                    sql = f"UPDATE client SET {', '.join(update_pairs)} WHERE customer_TIN = ?"

                cur.execute(sql, tuple(params))
                conn.commit()
                conn.close()
                messagebox.showinfo("Succès", "Client mis à jour", parent=dlg)
                dlg.destroy()
                if callable(refresh_cb): refresh_cb()
            except Exception as e:
                messagebox.showerror("Erreur", f"Échec mise à jour: {e}", parent=dlg)

        def _on_delete_client():
            # Defensive permission check
            if not _current_user_can_delete():
                messagebox.showerror("Permission", "Vous n'êtes pas autorisé à supprimer des clients.", parent=dlg)
                return
            if not messagebox.askyesno("Confirmer", "Supprimer ce client ?", parent=dlg):
                return
            try:
                conn = get_connection()
                cur = conn.cursor()
                cid = client.get("id")
                if cid is not None:
                    cur.execute("DELETE FROM client WHERE id = ?", (cid,))
                else:
                    cur.execute("DELETE FROM client WHERE customer_TIN = ?", (client.get("customer_TIN"),))
                conn.commit()
                conn.close()
                messagebox.showinfo("Supprimé", "Client supprimé ✅", parent=dlg)
                dlg.destroy()
                if callable(refresh_cb): refresh_cb()
            except Exception as e:
                messagebox.showerror("Erreur", f"Échec suppression: {e}", parent=dlg)

        btn_save = tk.Button(btn_frame, text="Enregistrer", bg="#007bff", fg="white", command=_validate_and_save)
        btn_save.pack(side="right", padx=6)

        # show or disable delete according to permission
        if _current_user_can_delete():
            btn_del = tk.Button(btn_frame, text="Supprimer", bg="#dc3545", fg="white", command=_on_delete_client)
        else:
            btn_del = tk.Button(btn_frame, text="Supprimer", bg="#dc3545", fg="white", state="disabled",
                                command=lambda: messagebox.showerror("Permission", "Vous n'êtes pas autorisé à supprimer des clients.", parent=dlg))
        btn_del.pack(side="right", padx=6)

        btn_cancel = tk.Button(btn_frame, text="Annuler", command=dlg.destroy)
        btn_cancel.pack(side="right", padx=6)

        dlg.wait_window()

    def _delete_client(nif, refresh_cb):
        if not _current_user_can_delete():
            messagebox.showerror("Permission", "Vous n'êtes pas autorisé à supprimer des clients.")
            return
        if not messagebox.askyesno("Confirmer", "Voulez-vous supprimer ce client ?"):
            return
        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("DELETE FROM client WHERE customer_TIN = ?", (nif,))
            conn.commit()
            conn.close()
            if callable(refresh_cb): refresh_cb()
            messagebox.showinfo("Supprimé", "Client supprimé ✅")
        except Exception as e:
            messagebox.showerror("Erreur", f"Échec suppression : {e}")

    created_row_widgets = []

    def clear_rows():
        for child in inner_grid.grid_slaves():
            info = child.grid_info()
            if int(info.get("row", 0)) >= 1:
                try: child.destroy()
                except Exception: pass
        created_row_widgets.clear()

    # Define refresh early so callbacks can safely reference it
    def refresh():
        render_rows(search_var.get().strip())

    def render_rows(filter_text=None):
        page = current_page["n"]
        page_size = page_size_var.get()
        try:
            rows, total = _fetch_clients_page(filter_text=filter_text, page=page, page_size=page_size)
            pagination_info["total"] = total
        except sqlite3.OperationalError as oe:
            clear_rows()
            lbl = tk.Label(inner_grid, text=f"Erreur base donnée: {oe}", bg=CARD_BG, fg="#900", font=("Segoe UI", 10, "bold"), anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            return
        except Exception as e:
            clear_rows()
            lbl = tk.Label(inner_grid, text=f"Erreur: {e}", bg=CARD_BG, fg="#900", font=("Segoe UI", 10, "bold"), anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            return

        clear_rows()

        if total == 0:
            lbl = tk.Label(inner_grid, text="Aucun client trouvé.", bg=CARD_BG, fg=LABEL_FG, font=("Segoe UI", 10), anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            return

        total_pages = max(1, (total + page_size - 1) // page_size)
        if current_page["n"] > total_pages:
            current_page["n"] = total_pages
            page = current_page["n"]
            rows, total = _fetch_clients_page(filter_text=filter_text, page=page, page_size=page_size)
            pagination_info["total"] = total

        for ri, row in enumerate(rows, start=1):
            bg = ROW_BG_1 if (ri % 2 == 1) else ROW_BG_2
            widgets = []
            for ci, (dbcol, _, _) in enumerate(COLUMNS):
                txt = _format_cell(dbcol, row[dbcol]) if dbcol in row.keys() else str(row[ci])
                lbl = tk.Label(inner_grid, text=txt, anchor="w", bg=bg, fg=LABEL_FG, font=FONT_CELL,
                                 padx=6, pady=6, bd=1, relief="solid", wraplength=150)
                lbl.grid(row=ri, column=ci, sticky="nsew", padx=0, pady=0)
                inner_grid.grid_rowconfigure(ri, minsize=ROW_HEIGHT)
                widgets.append(lbl)

            # Use id for actions
            client_id = row["id"] if "id" in row.keys() else row[0]
            act_frame = tk.Frame(inner_grid, bg=bg, bd=1, relief="solid")
            act_frame.grid(row=ri, column=len(COLUMNS), sticky="nsew", padx=0, pady=0)
            inner_grid.grid_columnconfigure(len(COLUMNS), minsize=200)

            btn_v = tk.Button(act_frame, text="🔍 Voir", bg="#6c757d", fg="white", activebackground="#5a6268", padx=6, pady=2,
                              command=lambda cid=client_id: _view_client_modal(_get_full_client_by_id(cid)))
            btn_e = tk.Button(act_frame, text="✏️ Editer", bg="#007bff", fg="white", activebackground="#0056b3", padx=6, pady=2,
                              command=lambda cid=client_id: _edit_client_modal(_get_full_client_by_id(cid), refresh))

            # deletion button depends on permission
            if _current_user_can_delete():
                btn_d = tk.Button(act_frame, text="🗑 Supprimer", bg="#dc3545", fg="white", activebackground="#c82333", padx=6, pady=2,
                                  command=lambda cid=client_id: _delete_client_by_id(cid, refresh, parent))
            else:
                btn_d = tk.Button(act_frame, text="🗑 Supprimer", bg="#dc3545", fg="white", activebackground="#c82333", padx=6, pady=2,
                                  state="disabled", command=lambda: messagebox.showerror("Permission", "Vous n'êtes pas autorisé à supprimer des clients.", parent=parent))

            btn_v.pack(side="left", padx=(4,0), pady=4)
            btn_e.pack(side="left", padx=4, pady=4)
            btn_d.pack(side="left", padx=(0,4), pady=4)

            widgets.append(act_frame)
            created_row_widgets.append(widgets)

        pager_text = f"Page {current_page['n']} / {total_pages} — {pagination_info['total']} client(s)"
        # Robust pager label creation/update (grid preferred)
        try:
            lbl = getattr(parent, "_pager_label", None)
            if lbl and getattr(lbl, "winfo_exists", lambda: False)():
                lbl.config(text=pager_text)
            else:
                parent._pager_label = tk.Label(parent, text=pager_text, bg=CONTENT_BG, fg=LABEL_FG, font=("Segoe UI", 9))
                try:
                    parent._pager_label.grid(row=2, column=0, sticky="w", padx=12, pady=(4,0))
                except tk.TclError:
                    try:
                        parent._pager_label.pack(side="left", anchor="w", padx=12, pady=(4,0))
                    except Exception:
                        pass
        except Exception:
            pass

        try:
            btn_prev.config(state="normal" if current_page["n"] > 1 else "disabled")
            btn_next.config(state="normal" if current_page["n"] < total_pages else "disabled")
        except Exception:
            pass

    # pagination controls
    pager_frame = tk.Frame(parent, bg=CONTENT_BG)
    pager_frame.grid(row=3, column=0, sticky="e", padx=12, pady=(6,12))
    btn_prev = tk.Button(pager_frame, text="◀ Préc", bg="#6c757d", fg=BTN_FG, width=8, activebackground="#5a6268")
    btn_next = tk.Button(pager_frame, text="Suiv ▶", bg="#6c757d", fg=BTN_FG, width=8, activebackground="#5a6268")
    btn_prev.pack(side="left", padx=6); btn_next.pack(side="left")

    def on_prev():
        if current_page["n"] > 1:
            current_page["n"] -= 1
            refresh()

    def on_next():
        page_size = page_size_var.get()
        total = pagination_info.get("total", 0)
        total_pages = max(1, (total + page_size - 1) // page_size)
        if current_page["n"] < total_pages:
            current_page["n"] += 1
            refresh()

    btn_prev.config(command=on_prev)
    btn_next.config(command=on_next)

    def on_page_size_change(e=None):
        current_page["n"] = 1
        refresh()

    page_size_cb.bind("<<ComboboxSelected>>", on_page_size_change)

    # ---------- Export functions (FR titles + centered PDF title) ----------
    def _export_all_rows():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM client ORDER BY customer_name")
        rows = cur.fetchall()
        desc = [d[0] for d in cur.description]
        conn.close()
        full_rows = [{desc[i]: row[i] for i in range(len(desc))} for row in rows]
        return full_rows, desc

    def _on_export_excel():
        try:
            import pandas as pd
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messagebox.showerror("Export", "Veuillez installer pandas et openpyxl. pip install pandas openpyxl")
            return

        rows, headers = _export_all_rows()
        if not rows:
            messagebox.showinfo("Export", "Aucune donnée à exporter.")
            return

        french_titles = {
            "id": "ID",
            "customer_name": "Nom",
            "customer_TIN": "NIF",
            "customer_type": "Type",
            "vat_customer_payer": "TVA",
        }
        df = pd.DataFrame(rows, columns=headers)
        df = df.rename(columns={h: french_titles.get(h, h) for h in headers})

        if "Type" in df.columns:
            df["Type"] = df["Type"].apply(lambda x: "Physique" if str(x) == "1" else "Morale" if str(x) == "2" else ("" if x is None else str(x)))
        if "TVA" in df.columns:
            df["TVA"] = df["TVA"].apply(lambda x: "Oui" if str(x) == "1" else "Non")

        fpath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Liste_Clients.xlsx", title="Exporter Excel")
        if not fpath:
            return

        try:
            with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Clients')
                ws = writer.sheets['Clients']
                header_fill = PatternFill(start_color="D9E6F6", end_color="D9E6F6", fill_type="solid")
                header_font = Font(bold=True)
                header_align = Alignment(horizontal='center', vertical='center')
                for cell in ws["1:1"]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                for col_idx, column in enumerate(ws.columns):
                    max_length = 0
                    column_letter = get_column_letter(col_idx + 1)
                    for cell in column:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = min(adjusted_width, 60)
            messagebox.showinfo("Export", "Export Excel terminé ✅")
        except Exception as e:
            messagebox.showerror("Erreur export", f"Échec de l'export Excel: {e}")

    def _on_export_pdf():
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            messagebox.showerror("Export", "Veuillez installer reportlab. pip install reportlab")
            return

        rows, headers = _export_all_rows()
        if not rows:
            messagebox.showinfo("Export", "Aucune donnée à exporter.")
            return

        french_titles = {
            "id": "ID",
            "customer_name": "Nom",
            "customer_TIN": "NIF",
            "customer_type": "Type",
            "vat_customer_payer": "TVA",
        }
        header_labels = [french_titles.get(h, h).replace("customer_", "").replace("_", " ").title() for h in headers]

        data = [header_labels]
        for r in rows:
            row_data = []
            for h in headers:
                val = r.get(h, "")
                if h == "customer_type":
                    val = "Physique" if str(val) == "1" else "Morale" if str(val) == "2" else ("" if val is None else str(val))
                elif h == "vat_customer_payer":
                    val = "Oui" if str(val) == "1" else "Non"
                row_data.append(str(val))
            data.append(row_data)

        fpath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile="Liste_Clients.pdf", title="Exporter PDF")
        if not fpath:
            return

        try:
            doc = SimpleDocTemplate(fpath, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                name="TitleCentered",
                parent=styles["Heading1"],
                alignment=TA_CENTER,
                fontSize=16,
                textColor=colors.HexColor("#0b3d91"),
                spaceAfter=12
            )
            title_para = Paragraph("Liste des clients", title_style)

            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#D9E6F6")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#212529")),
                ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#c7d2e7")),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("ALIGN", (0,0), (-1,0), "CENTER"),
                ("ALIGN", (0,1), (-1,-1), "LEFT"),
                ("LEFTPADDING", (0,0), (-1,-1), 3),
                ("RIGHTPADDING", (0,0), (-1,-1), 3),
            ]))

            elems = [title_para, Spacer(1, 12), table]
            doc.build(elems)
            messagebox.showinfo("Export", "Export PDF terminé ✅")
        except Exception as e:
            messagebox.showerror("Erreur export", f"Échec de l'export PDF: {e}")

    btn_export_xl.config(command=_on_export_excel)
    btn_export_pdf.config(command=_on_export_pdf)

    # live search binding
    def _on_search(*_):
        current_page["n"] = 1
        refresh()
    try:
        search_var.trace_add("write", _on_search)
    except Exception:
        try: search_var.trace("w", _on_search)
        except Exception: pass

    # initial render
    refresh()

    return {"refresh": lambda: render_rows(search_var.get().strip()),
            "search_var": search_var,
            "page_size_var": page_size_var}
