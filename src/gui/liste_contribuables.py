import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from database.connection import get_connection

def afficher_liste_contribuables(parent):
    """
    Vue des contribuables (table réduite à l'écran, exports complets).
    - Colonnes visibles dans la grille : id, tp_name, tp_TIN, tp_trade_number
    - Modals Voir / Modifier affichent tous les champs demandés (2 colonnes par ligne)
    - NIF (tp_TIN) non modifiable en édition
    - Exports Excel et PDF exportent toutes les colonnes listées dans ALL_FIELDS_KEYS
    - Textareas dans modals ont une taille réduite pour meilleure lisibilité (height=2 ou 3)
    Retourne dict {refresh, search_var, page_size_var}
    """
    # UI constants
    CONTENT_BG = "white"
    CARD_BG = "#ffffff"
    HEADER_FG = "#0b3d91"
    LABEL_FG = "#1f2937"
    ROW_BG_1 = CARD_BG
    ROW_BG_2 = "#fbfdff"
    FONT_TITLE = ("Segoe UI", 14, "bold")
    FONT_LABEL = ("Segoe UI", 10)
    FONT_CELL = ("Segoe UI", 10)
    BTN_VIEW_BG = "#6c757d"
    BTN_EDIT_BG = "#007bff"
    BTN_DELETE_BG = "#dc3545"

    # Visible columns in the on-screen table
    COLUMNS = [
        ("id", "ID", 60),
        ("tp_name", "Nom complet", 110),
        ("tp_TIN", "NIF", 120),
        ("tp_trade_number", "Numéro de registre", 140),
    ]

    # All columns to export and to show in modals (label, key, minwidth)
    ALL_FIELDS = [
        ("Nom complet", "tp_name"),
        ("NIF", "tp_TIN"),
        ("Numéro de registre", "tp_trade_number"),
        ("Numéro postal", "tp_postal_number"),
        ("Téléphone", "tp_phone_number"),
        ("Province", "tp_address_province"),
        ("Commune", "tp_address_commune"),
        ("Quartier", "tp_address_quartier"),
        ("Avenue", "tp_address_avenue"),
        ("Rue", "tp_address_rue"),
        ("Numéro de porte", "tp_address_number"),
        ("Centre fiscal", "tp_fiscal_center"),
        ("Forme juridique", "tp_legal_form"),
        ("Secteur d'activité", "tp_activity_sector"),
    ]
    ALL_FIELDS_KEYS = [k for _, k in ALL_FIELDS]

    PAGE_SIZES = [10, 15, 20, 50]
    page_size_var = tk.IntVar(value=15)
    current_page = {"n": 1}
    pagination_info = {"total": 0}
    
    # ----------------- DB helpers -----------------

    def _fetch_contribuables(filter_text="", page=1, page_size=15):
        q = f"%{filter_text}%" if filter_text else None
        conn = get_connection()
        cur = conn.cursor()
        cols = ", ".join([c[0] for c in COLUMNS])
        if q:
            cur.execute(f"""
                SELECT {cols}
                FROM contribuable
                WHERE tp_name LIKE ? OR tp_TIN LIKE ?
                ORDER BY id DESC
                LIMIT ? OFFSET ?
            """, (q, q, page_size, (page - 1) * page_size))
            cur2 = conn.cursor()
            cur2.execute("SELECT COUNT(1) FROM contribuable WHERE tp_name LIKE ? OR tp_TIN LIKE ?", (q, q))
            total = cur2.fetchone()[0]
        else:
            cur.execute(f"""
                SELECT {cols}
                FROM contribuable
                ORDER BY id DESC
                LIMIT ? OFFSET ?
            """, (page_size, (page - 1) * page_size))
            cur2 = conn.cursor()
            cur2.execute("SELECT COUNT(1) FROM contribuable")
            total = cur2.fetchone()[0]
        rows = cur.fetchall()
        conn.close()
        return rows, total

    def _get_contribuable_by_id(cid):
        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM contribuable WHERE id = ? LIMIT 1", (cid,))
            row = cur.fetchone()
            if not row:
                conn.close()
                return {}
            desc = [d[0] for d in cur.description]
            result = {desc[i]: row[i] for i in range(len(desc))}
            conn.close()
            return result
        except Exception:
            return {}

    def _delete_contribuable_by_id(cid):
        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("DELETE FROM contribuable WHERE id = ?", (cid,))
            conn.commit()
            conn.close()
            return True, None
        except Exception as e:
            return False, str(e)

    # ----------------- UI setup -----------------
    for w in parent.winfo_children():
        try: w.destroy()
        except Exception: pass

    parent.configure(bg=CONTENT_BG)
    parent.grid_columnconfigure(0, weight=1)
    parent.grid_rowconfigure(2, weight=1)

    # First logical line: title + search on same row
    title_row = tk.Frame(parent, bg=CONTENT_BG)
    title_row.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
    title_row.grid_columnconfigure(0, weight=1)
    tk.Label(title_row, text="📋 Liste des Contribuables", font=FONT_TITLE, bg=CONTENT_BG, fg=HEADER_FG).grid(row=0, column=0, sticky="w")

    right_search = tk.Frame(title_row, bg=CONTENT_BG)
    right_search.grid(row=0, column=1, sticky="e")
    tk.Label(right_search, text="Recherche :", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL).grid(row=0, column=0, sticky="w", padx=(0,6))
    search_var = tk.StringVar()
    search_entry = tk.Entry(right_search, textvariable=search_var, font=FONT_LABEL, width=30, bg="white", relief="solid", bd=1)
    search_entry.grid(row=0, column=1, sticky="w")

    # Second logical line: page size + export buttons
    controls = tk.Frame(parent, bg=CONTENT_BG)
    controls.grid(row=1, column=0, sticky="ew", padx=12, pady=(4,8))
    controls.grid_columnconfigure(0, weight=1)
    right = tk.Frame(controls, bg=CONTENT_BG)
    right.grid(row=0, column=1, sticky="e")
    tk.Label(right, text="Taille page:", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL).grid(row=0, column=0, sticky="w", padx=(0,6))
    page_size_cb = ttk.Combobox(right, values=PAGE_SIZES, textvariable=page_size_var, width=4, state="readonly")
    page_size_cb.grid(row=0, column=1, sticky="w", padx=(0,8))
    page_size_cb.set(page_size_var.get())
    btn_export_xl = tk.Button(right, text="Export Excel", bg="#16a34a", fg="white", activebackground="#15803d", width=12)
    btn_export_pdf = tk.Button(right, text="Export PDF", bg="#2563eb", fg="white", activebackground="#1e40af", width=12)
    btn_export_xl.grid(row=0, column=2, padx=(8,4))
    btn_export_pdf.grid(row=0, column=3, padx=(4,0))

    # Card area: table rendered with grid
    card = tk.Frame(parent, bg=CARD_BG)
    card.grid(row=2, column=0, sticky="nsew", padx=12, pady=6)
    card.grid_rowconfigure(0, weight=1)
    card.grid_columnconfigure(0, weight=1)

    inner_grid = tk.Frame(card, bg=CARD_BG)
    inner_grid.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
    inner_grid.grid_rowconfigure(0, weight=0)

    # Header columns
    for ci, (dbcol, title, minw) in enumerate(COLUMNS):
        lbl = tk.Label(inner_grid, text=title, bg="#eef6ff", fg=HEADER_FG,
                       font=("Segoe UI", 9, "bold"), anchor="w", padx=6, bd=1, relief="solid", pady=8)
        lbl.grid(row=0, column=ci, sticky="nsew", padx=0, pady=0)
        if dbcol == "tp_name":
            inner_grid.grid_columnconfigure(ci, weight=1, minsize=minw)
        else:
            inner_grid.grid_columnconfigure(ci, weight=0, minsize=minw)

    # Actions header
    act_col = len(COLUMNS)
    ha = tk.Label(inner_grid, text="Actions", bg="#eef6ff", fg=HEADER_FG,
                  font=("Segoe UI", 9, "bold"), anchor="center", padx=6, bd=1, relief="solid", pady=8)
    ha.grid(row=0, column=act_col, sticky="nsew", padx=0, pady=0)
    inner_grid.grid_columnconfigure(act_col, weight=0, minsize=220)

    # Pager / footer
    pager_frame = tk.Frame(parent, bg=CONTENT_BG)
    pager_frame.grid(row=3, column=0, sticky="ew", padx=12, pady=(6,12))
    pager_frame.grid_columnconfigure(0, weight=1)
    status_lbl = tk.Label(pager_frame, text="", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL)
    status_lbl.grid(row=0, column=0, sticky="w")
    btn_prev = tk.Button(pager_frame, text="◀ Préc", bg="#6c757d", fg="white", width=8)
    btn_next = tk.Button(pager_frame, text="Suiv ▶", bg="#6c757d", fg="white", width=8)
    btn_prev.grid(row=0, column=1, padx=6)
    btn_next.grid(row=0, column=2)

    created_row_widgets = []

    def clear_rows():
        for child in inner_grid.grid_slaves():
            info = child.grid_info()
            if int(info.get("row", 0)) >= 1:
                try: child.destroy()
                except Exception: pass
        created_row_widgets.clear()

    # ----------------- Modals -----------------
    def _center_window(win: tk.Toplevel, parent_widget: tk.Widget, prefer_w: int = None, prefer_h: int = None):
        try:
            parent_widget.update_idletasks()
            win.update_idletasks()
            px = parent_widget.winfo_rootx()
            py = parent_widget.winfo_rooty()
            pw = parent_widget.winfo_width()
            ph = parent_widget.winfo_height()
            req_w = win.winfo_reqwidth() or (prefer_w or 900)
            req_h = win.winfo_reqheight() or (prefer_h or 600)
            ww = prefer_w or req_w
            wh = prefer_h or req_h
            sw = parent_widget.winfo_screenwidth()
            sh = parent_widget.winfo_screenheight()
            ww = min(ww, sw - 40)
            wh = min(wh, sh - 40)
            try:
                max_w = max(300, pw - 40)
                max_h = max(200, ph - 40)
                if ww > max_w: ww = max_w
                if wh > max_h: wh = max_h
            except Exception:
                pass
            x = px + max(0, (pw - ww) // 2)
            y = py + max(0, (ph - wh) // 2)
            win.geometry(f"{ww}x{wh}+{x}+{y}")
        except Exception:
            try:
                sw = parent_widget.winfo_screenwidth()
                sh = parent_widget.winfo_screenheight()
                fw = prefer_w or 900
                fh = prefer_h or 600
                fx = max(0, (sw - fw) // 2)
                fy = max(0, (sh - fh) // 2)
                win.geometry(f"{fw}x{fh}+{fx}+{fy}")
            except Exception:
                pass

    def _view_modal(cid):
        data = _get_contribuable_by_id(cid)
        if not data:
            messagebox.showinfo("Détails contribuable", "Données introuvables.")
            return
        dlg = tk.Toplevel(parent)
        dlg.transient(parent); dlg.grab_set()
        dlg.title(f"🔍 Contribuable — {data.get('tp_name','')}")
        prefer_w, prefer_h = 920, 560
        _center_window(dlg, inner_grid, prefer_w=prefer_w, prefer_h=prefer_h)

        # scrollable area so everything remains visible
        canvas = tk.Canvas(dlg, bg=CARD_BG)
        vs = ttk.Scrollbar(dlg, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vs.set)
        body = tk.Frame(canvas, bg=CARD_BG)
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        def _on_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        body.bind("<Configure>", _on_configure)

        # two fields per row
        for i, (label_text, key) in enumerate(ALL_FIELDS):
            row = i // 2
            col = (i % 2) * 2
            tk.Label(body, text=label_text + " :", font=FONT_LABEL, anchor="w", bg=CARD_BG).grid(row=row, column=col, sticky="nw", padx=8, pady=(8,4))
            val = "" if data.get(key) is None else str(data.get(key))
            # smaller text widget (height 2 or 3)
            height = 1 if len(val) < 80 else 2
            widget = tk.Text(body, width=40, height=height, wrap="word", font=FONT_CELL)
            widget.grid(row=row, column=col+1, sticky="nw", padx=(6,12), pady=(8,4))
            widget.insert("1.0", val)
            widget.configure(state="disabled", bg="#f7f7f7")

        btn_close = tk.Button(dlg, text="Fermer", command=dlg.destroy, bg="#6c757d", fg="white")
        btn_close.pack(side="bottom", pady=8)

    def _edit_modal(cid):
        data = _get_contribuable_by_id(cid)
        if not data:
            messagebox.showinfo("Édition", "Contribuable introuvable.")
            return
        dlg = tk.Toplevel(parent)
        dlg.transient(parent); dlg.grab_set()
        dlg.title(f"✏️ Éditer contribuable — {data.get('tp_TIN','')}")
        prefer_w, prefer_h = 980, 640
        _center_window(dlg, inner_grid, prefer_w=prefer_w, prefer_h=prefer_h)

        canvas = tk.Canvas(dlg, bg=CARD_BG)
        vs = ttk.Scrollbar(dlg, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vs.set)
        body = tk.Frame(canvas, bg=CARD_BG)
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        def _on_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        body.bind("<Configure>", _on_configure)

        vars_map = {}
        # arrange two fields per row, columns: label/value | label/value
        for i, (label_text, key) in enumerate(ALL_FIELDS):
            row = i // 2
            col_offset = (i % 2) * 4
            tk.Label(body, text=label_text + " :", font=FONT_LABEL, bg=CARD_BG).grid(row=row, column=col_offset, sticky="w", padx=8, pady=6)
            v = "" if data.get(key) is None else str(data.get(key))
            if key == "tp_TIN":
                ent = tk.Entry(body, width=30, font=FONT_CELL)
                ent.grid(row=row, column=col_offset+1, sticky="w", padx=(6,12), pady=6)
                ent.insert(0, v)
                ent.configure(state="readonly", readonlybackground="#f0f0f0")
                vars_map[key] = ent
            else:
                # reduced-size textareas (height=2 or 3) for address/activity fields
                if len(v) > 120 or key.startswith("tp_address_") or key == "tp_activity_sector":
                    txt = tk.Text(body, width=36, height=2, font=FONT_CELL, wrap="word")
                    txt.grid(row=row, column=col_offset+1, sticky="w", padx=(6,12), pady=6)
                    txt.insert("1.0", v)
                    vars_map[key] = txt
                else:
                    sv = tk.StringVar(value=v)
                    ent = tk.Entry(body, textvariable=sv, width=36, font=FONT_CELL)
                    ent.grid(row=row, column=col_offset+1, sticky="w", padx=(6,12), pady=6)
                    vars_map[key] = sv

        # action buttons
        btn_frame = tk.Frame(dlg, bg=CARD_BG)
        btn_frame.pack(fill="x", pady=(8,12))
        def _save():
            name = (vars_map["tp_name"].get() if isinstance(vars_map["tp_name"], tk.StringVar)
                    else vars_map["tp_name"].get("1.0","end").strip() if hasattr(vars_map["tp_name"], "get") and not isinstance(vars_map["tp_name"], tk.Entry) else vars_map["tp_name"].get())
            if not name:
                messagebox.showwarning("Validation", "Le nom est requis.", parent=dlg); return
            # collect updates (skip tp_TIN)
            update_vals = {}
            for k, widget in vars_map.items():
                if k == "tp_TIN":
                    continue
                if isinstance(widget, tk.StringVar):
                    update_vals[k] = widget.get().strip()
                elif isinstance(widget, tk.Entry):
                    update_vals[k] = widget.get().strip()
                else:
                    update_vals[k] = widget.get("1.0","end").strip()
            try:
                conn = get_connection()
                cur = conn.cursor()
                cur.execute("PRAGMA table_info(contribuable)")
                cols_info = cur.fetchall()
                cols = [c["name"] if isinstance(c, sqlite3.Row) else c[1] for c in cols_info]
                pairs = []
                params = []
                for col_name, val in update_vals.items():
                    if col_name in cols:
                        pairs.append(f"{col_name} = ?")
                        params.append(val)
                if not pairs:
                    messagebox.showinfo("Mise à jour", "Aucune colonne à mettre à jour.", parent=dlg)
                    conn.close()
                    return
                params.append(cid)
                sql = f"UPDATE contribuable SET {', '.join(pairs)} WHERE id = ?"
                cur.execute(sql, tuple(params))
                conn.commit()
                conn.close()
                messagebox.showinfo("Succès", "Contribuable mis à jour.", parent=dlg)
                dlg.destroy()
                render_rows(search_var.get().strip())
            except Exception as e:
                messagebox.showerror("Erreur", f"Échec mise à jour : {e}", parent=dlg)

        tk.Button(btn_frame, text="Enregistrer", bg=BTN_EDIT_BG, fg="white", command=_save).pack(side="right", padx=8)
        tk.Button(btn_frame, text="Annuler", command=dlg.destroy).pack(side="right", padx=8)

    def _on_delete(cid):
        if not messagebox.askyesno("Supprimer", f"Confirmer la suppression du contribuable ID {cid} ?"):
            return
        ok, err = _delete_contribuable_by_id(cid)
        if not ok:
            messagebox.showerror("Erreur", f"Impossible de supprimer : {err}")
        else:
            render_rows(search_var.get().strip())
            messagebox.showinfo("Supprimé", "Contribuable supprimé avec succès.")

    # ----------------- Render rows -----------------
    def render_rows(filter_text=None):
        page = current_page["n"]
        page_size = page_size_var.get()
        try:
            rows, total = _fetch_contribuables(filter_text or "", page=page, page_size=page_size)
            pagination_info["total"] = total
        except sqlite3.OperationalError as oe:
            clear_rows()
            lbl = tk.Label(inner_grid, text=f"Erreur base donnée: {oe}", bg=CARD_BG, fg="#900", font=FONT_LABEL, anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            status_lbl.config(text="")
            return
        except Exception as e:
            clear_rows()
            lbl = tk.Label(inner_grid, text=f"Erreur: {e}", bg=CARD_BG, fg="#900", font=FONT_LABEL, anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            status_lbl.config(text="")
            return

        clear_rows()
        if total == 0:
            lbl = tk.Label(inner_grid, text="Aucun contribuable trouvé.", bg=CARD_BG, fg=LABEL_FG, font=FONT_LABEL, anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            status_lbl.config(text="0 contribuable(s) affiché(s)")
            return

        total_pages = max(1, (total + page_size - 1) // page_size)
        if current_page["n"] > total_pages:
            current_page["n"] = total_pages
            page = current_page["n"]
            rows, total = _fetch_contribuables(filter_text or "", page=page, page_size=page_size)
            pagination_info["total"] = total

        for ri, row in enumerate(rows, start=1):
            bg = ROW_BG_1 if (ri % 2 == 1) else ROW_BG_2
            for ci, (dbcol, _, _) in enumerate(COLUMNS):
                txt = row[dbcol] if dbcol in row.keys() else ""
                lbl = tk.Label(inner_grid, text=str(txt or ""), anchor="w", bg=bg, fg=LABEL_FG, font=FONT_CELL,
                               padx=6, pady=6, bd=1, relief="solid", wraplength=300)
                lbl.grid(row=ri, column=ci, sticky="nsew", padx=0, pady=0)
                inner_grid.grid_rowconfigure(ri, minsize=28)

            cid = row["id"]
            act_frame = tk.Frame(inner_grid, bg=bg, bd=1, relief="solid")
            act_frame.grid(row=ri, column=len(COLUMNS), sticky="nsew", padx=0, pady=0)
            inner_grid.grid_columnconfigure(len(COLUMNS), minsize=220)

            btn_v = tk.Button(act_frame, text="🔍 Voir", bg=BTN_VIEW_BG, fg="white", command=lambda idv=cid: _view_modal(idv))
            btn_e = tk.Button(act_frame, text="✏️ Modifier", bg=BTN_EDIT_BG, fg="white", command=lambda ide=cid: _edit_modal(ide))
            btn_d = tk.Button(act_frame, text="🗑 Supprimer", bg=BTN_DELETE_BG, fg="white", command=lambda idd=cid: _on_delete(idd))

            btn_v.pack(side="left", padx=(4,0), pady=4)
            btn_e.pack(side="left", padx=4, pady=4)
            btn_d.pack(side="left", padx=(0,4), pady=4)

            created_row_widgets.append([act_frame])

        pager_text = f"Page {current_page['n']} / {total_pages} — {pagination_info['total']} contribuable(s)"
        status_lbl.config(text=pager_text)
        btn_prev.config(state="normal" if current_page["n"] > 1 else "disabled")
        btn_next.config(state="normal" if current_page["n"] < total_pages else "disabled")

    # ----------------- Pagination handlers -----------------
    def on_prev():
        if current_page["n"] > 1:
            current_page["n"] -= 1
            render_rows(search_var.get().strip())

    def on_next():
        page_size = page_size_var.get()
        total = pagination_info.get("total", 0)
        total_pages = max(1, (total + page_size - 1) // page_size)
        if current_page["n"] < total_pages:
            current_page["n"] += 1
            render_rows(search_var.get().strip())

    btn_prev.config(command=on_prev)
    btn_next.config(command=on_next)
    def on_page_size_change(e=None):
        current_page["n"] = 1
        render_rows(search_var.get().strip())
    page_size_cb.bind("<<ComboboxSelected>>", on_page_size_change)

    # ----------------- Exports (export all ALL_FIELDS_KEYS) -----------------
    def _export_all_rows_full():
        conn = get_connection()
        cur = conn.cursor()
        cols = ", ".join(ALL_FIELDS_KEYS)
        cur.execute(f"SELECT {cols} FROM contribuable ORDER BY id DESC")
        rows = cur.fetchall()
        desc = [d[0] for d in cur.description]
        conn.close()
        full_rows = [{desc[i]: row[i] for i in range(len(desc))} for row in rows]
        return full_rows, desc

    def _on_export_excel():
        try:
            import pandas as pd
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messagebox.showerror("Export", "Veuillez installer pandas et openpyxl (pip install pandas openpyxl)")
            return
        rows, headers = _export_all_rows_full()
        if not rows:
            messagebox.showinfo("Export", "Aucune donnée à exporter.")
            return
        # mapping headers to french titles
        french_titles = {k: label for label, k in ALL_FIELDS}
        df = pd.DataFrame(rows, columns=headers)
        df = df.rename(columns={h: french_titles.get(h, h) for h in headers})
        fpath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Liste_contribuables.xlsx", title="Exporter Excel")
        if not fpath:
            return
        try:
            with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Contribuables')
                ws = writer.sheets['Contribuables']
                header_fill = PatternFill(start_color="D9E6F6", end_color="D9E6F6", fill_type="solid")
                header_font = Font(bold=True)
                header_align = Alignment(horizontal='center', vertical='center')
                for cell in ws["1:1"]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                for col_idx, column in enumerate(ws.columns):
                    max_length = 0
                    column_letter = get_column_letter(col_idx + 1)
                    for cell in column:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = min(adjusted_width, 80)
            messagebox.showinfo("Export", "Export Excel terminé ✅")
        except Exception as e:
            messagebox.showerror("Erreur export", f"Échec de l'export Excel: {e}")

    def _on_export_pdf():
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            messagebox.showerror("Export", "Veuillez installer reportlab (pip install reportlab)")
            return
        rows, headers = _export_all_rows_full()
        if not rows:
            messagebox.showinfo("Export", "Aucune donnée à exporter.")
            return
        french_titles = [label for label, key in ALL_FIELDS]
        data = [french_titles]
        for r in rows:
            data.append([str(r.get(h, "")) for h in headers])
        fpath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile="Liste_contribuables.pdf", title="Exporter PDF")
        if not fpath:
            return
        try:
            doc = SimpleDocTemplate(fpath, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(name="TitleCentered", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=14, textColor=colors.HexColor("#0b3d91"))
            title_para = Paragraph("Liste des Contribuables", title_style)
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#D9E6F6")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#212529")),
                ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#c7d2e7")),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("ALIGN", (0,0), (-1,0), "CENTER"),
                ("ALIGN", (0,1), (-1,-1), "LEFT"),
                ("LEFTPADDING", (0,0), (-1,-1), 3),
                ("RIGHTPADDING", (0,0), (-1,-1), 3),
            ]))
            elems = [title_para, Spacer(1,12), table]
            doc.build(elems)
            messagebox.showinfo("Export", "Export PDF terminé ✅")
        except Exception as e:
            messagebox.showerror("Erreur export", f"Échec de l'export PDF: {e}")

    btn_export_xl.config(command=_on_export_excel)
    btn_export_pdf.config(command=_on_export_pdf)

    # ----------------- Bindings -----------------
    def _on_search_event(e=None):
        current_page["n"] = 1
        render_rows(search_var.get().strip())
    search_entry.bind("<KeyRelease>", _on_search_event)
    search_entry.bind("<Return>", _on_search_event)
    try:
        search_var.trace_add("write", lambda *a: _on_search_event())
    except Exception:
        try: search_var.trace("w", lambda *a: _on_search_event())
        except Exception: pass

    # initial render
    render_rows = locals().get('render_rows')  # placeholder access; real function defined below
    # define render_rows then call it
    def render_rows_wrapper(ft=""):
        return render_rows(ft) if callable(render_rows) else None
    # actual render_rows defined earlier in closure — ensure it's present
    # (the real render_rows is defined above; call it now)
    # call the render function that's in this scope
    try:
        # if render_rows is defined as a real function in locals, use it
        if callable(locals().get('render_rows')):
            locals()['render_rows'](search_var.get().strip())
        else:
            # fallback: call the nested render_rows function created earlier by this definition
            # (in this file render_rows is defined later, so simply call the function name used in this closure)
            # to avoid name errors call the inner render routine explicitly
            pass
    except Exception:
        # safe call: call the function defined in this scope (we have one defined below)
        try:
            render_rows(search_var.get().strip())
        except Exception:
            pass

    # The actual render_rows function (re-defined here to ensure closure correctness)
    def render_rows(filter_text=None):
        # implementation already present above; redirect to that one via the closure.
        # replicate minimal rendering logic by calling the internal fetch + building rows
        page = current_page["n"]
        page_size = page_size_var.get()
        try:
            rows, total = _fetch_contribuables(filter_text or "", page=page, page_size=page_size)
            pagination_info["total"] = total
        except Exception as e:
            clear_rows()
            lbl = tk.Label(inner_grid, text=f"Erreur: {e}", bg=CARD_BG, fg="#900", font=FONT_LABEL, anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            status_lbl.config(text="")
            return

        clear_rows()
        if total == 0:
            lbl = tk.Label(inner_grid, text="Aucun contribuable trouvé.", bg=CARD_BG, fg=LABEL_FG, font=FONT_LABEL, anchor="w")
            lbl.grid(row=1, column=0, columnspan=len(COLUMNS)+1, sticky="ew", padx=8, pady=8)
            created_row_widgets.append([lbl])
            status_lbl.config(text="0 contribuable(s) affiché(s)")
            return

        total_pages = max(1, (total + page_size - 1) // page_size)
        if current_page["n"] > total_pages:
            current_page["n"] = total_pages
            page = current_page["n"]
            rows, total = _fetch_contribuables(filter_text or "", page=page, page_size=page_size)
            pagination_info["total"] = total

        for ri, row in enumerate(rows, start=1):
            bg = ROW_BG_1 if (ri % 2 == 1) else ROW_BG_2
            for ci, (dbcol, _, _) in enumerate(COLUMNS):
                txt = row[dbcol] if dbcol in row.keys() else ""
                lbl = tk.Label(inner_grid, text=str(txt or ""), anchor="w", bg=bg, fg=LABEL_FG, font=FONT_CELL,
                               padx=6, pady=6, bd=1, relief="solid", wraplength=300)
                lbl.grid(row=ri, column=ci, sticky="nsew", padx=0, pady=0)
                inner_grid.grid_rowconfigure(ri, minsize=28)
            cid = row["id"]
            act_frame = tk.Frame(inner_grid, bg=bg, bd=1, relief="solid")
            act_frame.grid(row=ri, column=len(COLUMNS), sticky="nsew", padx=0, pady=0)
            inner_grid.grid_columnconfigure(len(COLUMNS), minsize=220)
            btn_v = tk.Button(act_frame, text="🔍 Voir", bg=BTN_VIEW_BG, fg="white", command=lambda idv=cid: _view_modal(idv))
            btn_e = tk.Button(act_frame, text="✏️ Modifier", bg=BTN_EDIT_BG, fg="white", command=lambda ide=cid: _edit_modal(ide))
            btn_d = tk.Button(act_frame, text="🗑 Supprimer", bg=BTN_DELETE_BG, fg="white", command=lambda idd=cid: _on_delete(idd))
            btn_v.pack(side="left", padx=(4,0), pady=4)
            btn_e.pack(side="left", padx=4, pady=4)
            btn_d.pack(side="left", padx=(0,4), pady=4)
            created_row_widgets.append([act_frame])

        pager_text = f"Page {current_page['n']} / {total_pages} — {pagination_info['total']} contribuable(s)"
        status_lbl.config(text=pager_text)
        btn_prev.config(state="normal" if current_page["n"] > 1 else "disabled")
        btn_next.config(state="normal" if current_page["n"] < total_pages else "disabled")

    # replace the local render_rows reference with the real one
    locals()['render_rows'] = render_rows

    # initial call
    render_rows(search_var.get().strip())

    return {"refresh": lambda: render_rows(search_var.get().strip()),
            "search_var": search_var,
            "page_size_var": page_size_var}
