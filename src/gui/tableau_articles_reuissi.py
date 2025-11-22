# gestion_obr_articles_viewer_with_cost_compact_actions_single.py
# -*- coding: utf-8 -*-
"""
Viewer compact — lecture depuis la table mouvement_stock avec LEFT JOIN article_stock_local,
affichage compact, exports Excel/PDF via asksaveasfilename, résumé et détail par ligne.

- Lecture: SELECT {select_expr} FROM mouvement_stock ms
         LEFT JOIN article_stock_local a ON ms.article_stock_id = a.id
         WHERE ms.obr_status = 1
- Colonne Actions contient un unique bouton "Voir" par ligne.
- Exports excluent la colonne UI-only Actions.
- Date filters, movement type filter, pagination simple.
"""
from __future__ import annotations
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import os
import logging
import math
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from typing import List, Dict, Any, Optional, Tuple

try:
    from tkcalendar import DateEntry
    _HAS_DATEENTRY = True
except Exception:
    DateEntry = None
    _HAS_DATEENTRY = False

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

from database.connection import get_connection

CONTENT_BG = "#f6f8fa"
CARD_BG = "#ffffff"
CONTOUR_BG = "#e6eef9"
TITLE_FG = "#0b3d91"
LABEL_FG = "#1f2937"
ROW_ALT = "#fbfdff"
FONT_TITLE = ("Segoe UI", 16, "bold")
FONT_LABEL = ("Segoe UI", 11)
FONT_CELL = ("Segoe UI", 11)

MOVEMENT_CHOICES = [
    ("Entrée normale (EN)", "EN"), ("Entrée retour (ER)", "ER"), ("Entrée initiale (EI)", "EI"),
    ("Entrée ajustement (EAJ)", "EAJ"), ("Entrée transfert (ET)", "ET"), ("Entrée autre unité (EAU)", "EAU"),
    ("Sortie normale (SN)", "SN"), ("Sortie perte (SP)", "SP"), ("Vente (SV)", "SV"),
    ("Sortie don (SD)", "SD"), ("Sortie consommation (SC)", "SC"), ("Sortie ajustement (SAJ)", "SAJ"),
    ("Sortie transfert (ST)", "ST"), ("Sortie autre unité (SAU)", "SAU"),
]

MOVEMENT_TYPE_DISPLAY = {name: code for name, code in MOVEMENT_CHOICES}
for _, code in MOVEMENT_CHOICES:
    MOVEMENT_TYPE_DISPLAY[code] = code

PAGE_SIZE = 10

# Table display columns (Actions included as a regular column at the end)
TABLE_COLS = [
    ("item_movement_date", "Date", 15),
    ("movement_type", "Type", 8),
    ("item_code", "Code article", 12),
    ("item_designation", "Désignation", 18),
    ("item_quantity", "Qté", 8),
    ("item_price", "PV unitaire", 10),
    ("item_cost_price", "PA unitaire", 10),
    ("actions", "Actions", 8)
]

# ALL_COLS selects from mouvement_stock (ms) and article_stock_local (a) where needed
ALL_COLS = [
    ("ms.id", "id"),
    ("ms.item_movement_date", "item_movement_date"),
    ("ms.item_movement_invoice_ref", "réf_facture"),
    ("ms.item_movement_description", "description_mouvement"),
    ("ms.item_movement_type", "movement_type"),
    ("ms.item_code", "item_code"),
    ("ms.item_designation", "item_designation"),
    ("ms.item_quantity", "item_quantity"),
    # Prefer article sale/cost if available, fallback to mouvement_stock purchase_or_sale
    ("COALESCE(a.item_sale_price, ms.item_purchase_or_sale_price)", "item_price"),
    ("COALESCE(a.item_cost_price, ms.item_purchase_or_sale_price)", "item_cost_price"),
    ("a.item_ct", "item_ct"),
    ("a.item_tl", "item_tl"),
    ("a.item_tsce_tax", "item_tsce_tax"),
    ("a.item_ott_tax", "item_ott_tax"),
    ("ms.item_measurement_unit", "item_measurement_unit"),
]

LABELS_FR = {
    "item_code": "Code article",
    "item_designation": "Désignation",
    "item_movement_date": "Date mouvement",
    "réf_facture": "Réf facture",
    "description_mouvement": "Description",
    "item_quantity": "Quantité",
    "item_price": "Prix unitaire",
    "item_cost_price": "Prix achat",
    "item_ct": "CT",
    "item_tl": "TL",
    "item_tsce_tax": "TSCE",
    "item_ott_tax": "OTT",
    "movement_type": "Type",
    "actions": "Actions"
}

def query_declared_articles(date_from: Optional[str] = None, date_to: Optional[str] = None, movement_type_filter: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Lit les enregistrements déclarés (ms.obr_status = 1) en joignant article_stock_local.
    Retourne une liste de dicts avec les alias définis dans ALL_COLS.
    """
    conn = get_connection()
    cur = conn.cursor()
    select_expr = ", ".join([f"{sel} AS {alias}" for sel, alias in ALL_COLS])
    q = f"""
      SELECT {select_expr}
      FROM mouvement_stock ms
      LEFT JOIN article_stock_local a ON ms.article_stock_id = a.id
      WHERE ms.obr_status = 1
    """
    params: List[Any] = []
    if movement_type_filter and movement_type_filter != "ALL":
        q += " AND ms.item_movement_type = ?"
        params.append(movement_type_filter)
    else:
        q += " AND (ms.item_movement_type IN ('EN','ER','EI','EAJ','ET','EAU') OR ms.item_movement_type IS NOT NULL)"
    if date_from:
        q += " AND ms.item_movement_date >= ?"; params.append(date_from)
    if date_to:
        q += " AND ms.item_movement_date <= ?"; params.append(date_to)
    q += " ORDER BY ms.item_movement_date DESC"
    cur.execute(q, params)
    rows = cur.fetchall()
    conn.close()
    col_names = [alias for _, alias in ALL_COLS]
    # Map rows to dicts
    results = [dict(zip(col_names, r)) for r in rows]
    return results

def export_to_excel_pandas(data: List[Dict[str, Any]], filename: str) -> Tuple[bool, Optional[str]]:
    try:
        df = pd.DataFrame(data)
        excluded = ("id",)
        order = [alias for _, alias in ALL_COLS if alias in df.columns and alias not in excluded]
        if "actions" in order:
            order.remove("actions")
        if not order:
            return False, "Aucune colonne à exporter"
        df = df[order]
        if "réf_facture" in df.columns:
            df["réf_facture"] = df["réf_facture"].apply(lambda v: "-" if (v is None or str(v).strip() == "") else v)
        df.columns = [LABELS_FR.get(col, col) for col in df.columns]
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Articles_Déclarés")
            ws = writer.book["Articles_Déclarés"]
            header_font = Font(bold=True, size=12)
            for col_idx, col in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx); cell.font = header_font
                max_len = max(ws.cell(row=r, column=col_idx).value and len(str(ws.cell(row=r, column=col_idx).value)) or 0 for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, len(col) + 2), 60)
            fill = PatternFill(start_color="D9E6F6", end_color="D9E6F6", fill_type="solid")
            thin = Side(border_style="thin", color="BDBDBD")
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = fill
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in r:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.font = Font(size=11)
        return True, None
    except Exception as e:
        logger.exception("Export Excel failed: %s", e)
        return False, str(e)

def export_to_pdf_reportlab(data: List[Dict[str, Any]], filename: str, title: str = "Articles déclarés OBR") -> Tuple[bool, Optional[str]]:
    try:
        header_aliases = [alias for _, alias in ALL_COLS if alias not in ("id", "movement_type")]
        if "actions" in header_aliases:
            header_aliases.remove("actions")
        header_labels = [LABELS_FR.get(alias, alias) for alias in header_aliases]
        table_data: List[List[str]] = [header_labels]
        for row in data:
            row_vals: List[str] = []
            for alias in header_aliases:
                val = row.get(alias, "")
                if alias == "réf_facture" and (val is None or str(val).strip() == ""):
                    val = "-"
                if alias == "item_movement_date" and val:
                    try:
                        val = datetime.strptime(val, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        pass
                if alias == "item_quantity":
                    qty = row.get("item_quantity", "")
                    if qty is None:
                        qty = ""
                    try:
                        qf = float(qty) if qty not in (None, "") else None
                        if qf is None:
                            val = ""
                        else:
                            val = f"{qf:.2f}"
                    except Exception:
                        val = str(qty)
                row_vals.append(str(val))
            table_data.append(row_vals)

        page_w, _ = landscape(A4)
        usable_w = page_w - 36
        alias_order = header_aliases
        col_widths: List[float] = []
        for alias in alias_order:
            if alias == "item_movement_date":
                col_widths.append(usable_w * 0.12)
            elif alias == "réf_facture":
                col_widths.append(usable_w * 0.10)
            elif alias == "description_mouvement":
                col_widths.append(usable_w * 0.14)
            elif alias == "item_code":
                col_widths.append(usable_w * 0.08)
            elif alias == "item_designation":
                col_widths.append(usable_w * 0.18)
            elif alias == "item_quantity":
                col_widths.append(usable_w * 0.08)
            elif alias in ("item_price", "item_cost_price"):
                col_widths.append(usable_w * 0.06)
            elif alias in ("item_ct", "item_tl", "item_tsce_tax", "item_ott_tax"):
                col_widths.append(usable_w * 0.03)
            else:
                col_widths.append(usable_w * (1.0 / len(alias_order)))
        s = sum(col_widths)
        col_widths = [w * (usable_w / s) for w in col_widths]

        doc = SimpleDocTemplate(filename, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", parent=styles["Heading1"], alignment=0, fontSize=16, leading=20, textColor=colors.HexColor("#0b3d91"))
        elems = [Paragraph(title, title_style), Spacer(1, 8)]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#D9E6F6")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#0b3d91")),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("ALIGN",(0,0),(-1,0),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#c7d2e7")),
            ("LEFTPADDING",(0,0),(-1,-1),3),
            ("RIGHTPADDING",(0,0),(-1,-1),3),
        ]))
        elems.append(tbl)
        doc.build(elems)
        return True, None
    except Exception as e:
        logger.exception("Export PDF failed: %s", e)
        return False, str(e)

def parse_date_input(s: Optional[str]) -> Optional[str]:
    s = s.strip() if s else ""
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None

def format_date_short(txt: Optional[str]) -> str:
    if not txt:
        return ""
    try:
        dt = datetime.strptime(txt, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(txt)

def _show_row_details(root, row: Dict[str, Any]):
    top = tk.Toplevel(root)
    top.title("Détail")
    top.geometry("820x480")
    top.resizable(False, False)

    outer = tk.Frame(top, bg=CONTOUR_BG, padx=8, pady=8)
    outer.pack(fill="both", expand=True, padx=8, pady=8)
    panel = tk.Frame(outer, bg=CARD_BG, bd=1, relief="solid", padx=10, pady=10)
    panel.pack(fill="both", expand=True)

    title_lbl = tk.Label(panel, text="Détail de la ligne", font=("Segoe UI", 13, "bold"), bg=CARD_BG, fg=TITLE_FG)
    title_lbl.grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 10))

    display_fields = [
        (alias, LABELS_FR.get(alias, alias.replace("_", " ").capitalize()))
        for _, alias in ALL_COLS
        if alias not in ("id", "movement_type")
    ]

    pairs_per_row = 2
    total_pairs = len(display_fields)
    rows_needed = math.ceil(total_pairs / pairs_per_row)

    idx = 0
    value_wrap = 160
    for row_idx in range(rows_needed):
        for col_block in range(pairs_per_row):
            if idx >= total_pairs:
                empty_lbl = tk.Label(panel, text="", bg=CARD_BG)
                empty_lbl.grid(row=1 + row_idx, column=col_block * 2, sticky="nsew", padx=6, pady=6)
                empty_val = tk.Label(panel, text="", bg=CARD_BG)
                empty_val.grid(row=1 + row_idx, column=col_block * 2 + 1, sticky="nsew", padx=6, pady=6)
                continue
            alias, human_label = display_fields[idx]
            lbl = tk.Label(panel, text=human_label + " :", font=("Segoe UI",11,"bold"), bg=CARD_BG, fg=LABEL_FG, anchor="e")
            lbl.grid(row=1 + row_idx, column=col_block * 2, sticky="e", padx=(4,8), pady=6)
            raw_val = row.get(alias, "")
            if alias == "item_movement_date" and raw_val:
                try: raw_val = datetime.strptime(raw_val, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M")
                except Exception: pass
            if alias == "réf_facture" and (raw_val is None or str(raw_val).strip() == ""):
                raw_val = "-"
            if alias == "item_quantity":
                unit = row.get("item_measurement_unit", "")
                try:
                    qf = float(raw_val) if raw_val not in (None, "") else None
                    if qf is None:
                        raw_val = ""
                    else:
                        raw_val = f"{qf:.2f}"
                        if unit:
                            raw_val = f"{raw_val} {unit}"
                except Exception:
                    raw_val = str(raw_val)
            val_frame = tk.Frame(panel, bg=CARD_BG, bd=1, relief="solid")
            val_frame.grid(row=1 + row_idx, column=col_block * 2 + 1, sticky="nsew", padx=(0,8), pady=6)
            val_lbl = tk.Label(val_frame, text=str(raw_val), font=("Segoe UI",9), bg=CARD_BG, fg=LABEL_FG, justify="left", wraplength=value_wrap, anchor="w")
            val_lbl.pack(fill="both", expand=True, padx=8, pady=8)
            panel.grid_columnconfigure(col_block*2, weight=0, minsize=110)
            panel.grid_columnconfigure(col_block*2+1, weight=1, minsize=140)
            idx += 1

    btn_frame = tk.Frame(panel, bg=CARD_BG)
    btn_frame.grid(row=1 + rows_needed, column=0, columnspan=pairs_per_row * 2, sticky="e", pady=(8,0))
    copy_btn = tk.Button(btn_frame, text="Copier JSON", bg="#10b981", fg="white", activebackground="#059669", padx=10, pady=6)
    def _copy_json():
        try:
            payload = {alias: row.get(alias, "") for _, alias in ALL_COLS}
            top.clipboard_clear()
            top.clipboard_append(json.dumps(payload, ensure_ascii=False))
            messagebox.showinfo("Copié","JSON copié dans le presse-papiers.", parent=top)
        except Exception:
            messagebox.showwarning("Erreur","Impossible de copier le JSON.", parent=top)
    copy_btn.config(command=_copy_json)
    copy_btn.pack(side="right", padx=(8,0))
    close_btn = tk.Button(btn_frame, text="Fermer", bg="#6b7280", fg="white", activebackground="#4b5563", padx=10, pady=6, command=top.destroy)
    close_btn.pack(side="right", padx=(8,0))

def show_obr_articles(parent):
    for w in parent.winfo_children():
        w.destroy()
    try:
        parent.configure(bg=CONTENT_BG)
    except Exception:
        pass

    state: Dict[str, Any] = {"data": [], "page": 1, "page_size": PAGE_SIZE, "total_pages": 1, "date_from": None, "date_to": None, "movement_type": "ALL"}

    title_frame = tk.Frame(parent, bg=CONTENT_BG)
    title_frame.pack(fill="x", padx=12, pady=(12, 6))
    tk.Label(title_frame, text="Tableau des articles déclarés à l'OBR réuissis", font=FONT_TITLE, bg=CONTENT_BG, fg=TITLE_FG).pack(anchor="w")
    tk.Label(title_frame, text="Sélectionner les dates et le type puis cliquer sur Recherche", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).pack(anchor="w", pady=(2, 8))

    ctrl = tk.Frame(parent, bg=CONTENT_BG)
    ctrl.pack(fill="x", padx=12, pady=(0, 8))

    row1 = tk.Frame(ctrl, bg=CONTENT_BG)
    row1.grid(row=0, column=0, sticky="w", padx=0, pady=(0, 6))

    tk.Label(row1, text="Date de :", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).grid(row=0, column=0, sticky="w", padx=(0,6))
    date_from_var = tk.StringVar()
    if _HAS_DATEENTRY:
        e_from = DateEntry(row1, textvariable=date_from_var, date_pattern="yyyy-mm-dd", width=14)
    else:
        e_from = ttk.Entry(row1, textvariable=date_from_var, width=20)
    e_from.grid(row=0, column=1, padx=(0,12))

    tk.Label(row1, text="Date à :", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).grid(row=0, column=2, sticky="w", padx=(0,6))
    date_to_var = tk.StringVar()
    if _HAS_DATEENTRY:
        e_to = DateEntry(row1, textvariable=date_to_var, date_pattern="yyyy-mm-dd", width=14)
    else:
        e_to = ttk.Entry(row1, textvariable=date_to_var, width=20)
    e_to.grid(row=0, column=3, padx=(0,12))

    tk.Label(row1, text="Type mouvement :", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).grid(row=0, column=4, sticky="w", padx=(0,6))
    movement_type_var = tk.StringVar(value="ALL")
    movement_options = ["ALL"] + [code for _, code in MOVEMENT_CHOICES]
    cb_type = ttk.Combobox(row1, textvariable=movement_type_var, values=movement_options, width=10, state="readonly")
    cb_type.grid(row=0, column=5, padx=(0,12))
    cb_type.set("ALL")

    row2 = tk.Frame(ctrl, bg=CONTENT_BG)
    row2.grid(row=1, column=0, sticky="w")

    btn_search = tk.Button(row2, text="Recherche", bg="#2563eb", fg="white", activebackground="#1e40af", padx=10)
    btn_search.grid(row=0, column=0, padx=(0,6))
    btn_refresh = tk.Button(row2, text="Rafraîchir", bg="#9ca3af", fg="white", activebackground="#6b7280", padx=10)
    btn_refresh.grid(row=0, column=1, padx=(0,6))
    btn_export_xls = tk.Button(row2, text="Export Excel", bg="#16a34a", fg="white", activebackground="#15803d", padx=10)
    btn_export_xls.grid(row=0, column=2, padx=(0,6))
    btn_export_pdf = tk.Button(row2, text="Export PDF", bg="#16a34a", fg="white", activebackground="#15803d", padx=10)
    btn_export_pdf.grid(row=0, column=3, padx=(0,6))

    card = tk.Frame(parent, bg=CARD_BG)
    card.pack(fill="both", padx=12, pady=(0,8), expand=True)
    inner_outer = tk.Frame(card, bg=CONTOUR_BG, bd=1, relief="solid")
    inner_outer.pack(fill="both", expand=True, padx=8, pady=8)
    inner_outer.grid_rowconfigure(0, weight=1)
    inner_outer.grid_columnconfigure(0, weight=1)
    inner_outer.grid_columnconfigure(1, weight=0)
    inner_outer.grid_columnconfigure(2, weight=1)

    inner = tk.Frame(inner_outer, bg=CARD_BG)
    inner.grid(row=0, column=1, sticky="nsew", padx=0, pady=8)
    inner.grid_rowconfigure(0, weight=0)

    header_font = ("Segoe UI",11,"bold")
    header_widgets = []
    for c, col in enumerate(TABLE_COLS):
        key, label, w = col
        hdr = tk.Label(inner, text=label, bg="#eef6ff", fg=LABEL_FG, font=header_font, anchor="w", bd=1, relief="solid", padx=8, pady=8)
        hdr.grid(row=0, column=c, sticky="nsew", padx=0, pady=0)
        header_widgets.append(hdr)

    for c, col in enumerate(TABLE_COLS):
        key, label, w = col
        base_width = int(w * 7)
        if label == "Date":
            inner.grid_columnconfigure(c, weight=0, minsize=115)
        elif label == "Désignation":
            inner.grid_columnconfigure(c, weight=1, minsize=120)
        elif label == "Code article":
            inner.grid_columnconfigure(c, weight=0, minsize=80)
        elif label == "Type":
            inner.grid_columnconfigure(c, weight=0, minsize=50)
        elif label in ("PV unitaire", "PA unitaire"):
            inner.grid_columnconfigure(c, weight=0, minsize=75)
        elif label == "Qté":
            inner.grid_columnconfigure(c, weight=0, minsize=60)
        elif label == "Actions":
            inner.grid_columnconfigure(c, weight=0, minsize=90)
        else:
            inner.grid_columnconfigure(c, weight=0, minsize=base_width)

    pager_frame = tk.Frame(parent, bg=CONTENT_BG)
    pager_frame.pack(fill="x", padx=12, pady=(0,12))
    lbl_page_info = tk.Label(pager_frame, text="", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL)
    lbl_page_info.pack(side="right", padx=(8,12))
    btn_prev = tk.Button(pager_frame, text="Précédent", state="disabled", padx=8)
    btn_prev.pack(side="left", padx=(0,8))
    btn_next = tk.Button(pager_frame, text="Suivant", state="disabled", padx=8)
    btn_next.pack(side="left", padx=(0,8))

    row_widgets: List[List[tk.Widget]] = []

    def load_data_and_refresh(page: int = 1):
        state["date_from"] = parse_date_input(date_from_var.get())
        state["date_to"] = parse_date_input(date_to_var.get())
        state["movement_type"] = movement_type_var.get() or "ALL"
        try:
            state["data"] = query_declared_articles(date_from=state["date_from"], date_to=state["date_to"], movement_type_filter=state["movement_type"])
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de charger les données: {e}", parent=parent)
            state["data"] = []
        total = len(state["data"])
        state["page_size"] = PAGE_SIZE
        state["total_pages"] = max(1, math.ceil(total / state["page_size"]))
        state["page"] = max(1, min(page, state["total_pages"]))
        refresh_table()

    def refresh_table():
        for widgets_row in row_widgets:
            for w in widgets_row:
                try: w.destroy()
                except Exception: pass
        row_widgets.clear()

        for child in inner.grid_slaves():
            info = child.grid_info()
            if int(info.get("row", 0)) >= 1:
                try: child.destroy()
                except Exception: pass

        data = state["data"]
        if not data:
            lbl_empty = tk.Label(inner, text="Aucun enregistrement pour les filtres choisis.", font=("Segoe UI",12), bg=CARD_BG, fg=LABEL_FG, padx=20, pady=20)
            lbl_empty.grid(row=1, column=0, columnspan=len(TABLE_COLS), sticky="nsew")
            lbl_page_info.config(text="Page 0 / 0")
            btn_prev.config(state="disabled"); btn_next.config(state="disabled")
            return

        page = state["page"]; size = state["page_size"]
        start = (page-1)*size; end = start+size
        page_rows = data[start:end]

        for r_index, row in enumerate(page_rows, start=1):
            widgets_row: List[tk.Widget] = []
            bg = ROW_ALT if (r_index % 2 == 0) else CARD_BG
            for c, col in enumerate(TABLE_COLS):
                alias = col[0]
                if alias == "actions":
                    action_frame = tk.Frame(inner, bg=bg, bd=1, relief="solid", padx=4, pady=4)
                    action_frame.grid(row=r_index, column=c, sticky="nsew", padx=0, pady=0)
                    def make_view_fn(row_data=row):
                        def _view(): _show_row_details(parent, row_data)
                        return _view
                    view_btn = tk.Button(action_frame, text="Voir", bg="#2563eb", fg="white", activebackground="#1e40af",
                                         padx=6, pady=2, command=make_view_fn())
                    view_btn.pack(side="left", padx=(4,4))
                    widgets_row.append(action_frame)
                    continue

                val = row.get(alias, "")
                if alias == "item_movement_date":
                    val = format_date_short(val)
                elif alias == "movement_type":
                    val = MOVEMENT_TYPE_DISPLAY.get(val, val)
                elif alias in ("item_price", "item_cost_price"):
                    try:
                        valf = float(val) if val is not None and str(val).strip() != "" else None
                        val = f"{valf:.2f}" if valf is not None else ""
                    except Exception:
                        val = str(val)
                elif alias == "item_quantity":
                    try:
                        qf = float(val) if val is not None and str(val).strip() != "" else None
                        if qf is None:
                            val = ""
                        else:
                            val = f"{qf:.2f}"
                            unit = row.get("item_measurement_unit", "")
                            if unit:
                                val = f"{val} {unit}"
                    except Exception:
                        val = str(val)
                cell = tk.Label(inner, text=str(val), bg=bg, fg=LABEL_FG, font=FONT_CELL, anchor="w", bd=1, relief="solid", padx=6, pady=6, wraplength=180)
                cell.grid(row=r_index, column=c, sticky="nsew", padx=0, pady=0)
                widgets_row.append(cell)
            row_widgets.append(widgets_row)

        lbl_page_info.config(text=f"Page {state['page']} / {state['total_pages']}")
        btn_prev.config(state="normal" if state["page"]>1 else "disabled")
        btn_next.config(state="normal" if state["page"]<state["total_pages"] else "disabled")

    def go_prev():
        if state["page"]>1:
            state["page"]-=1; refresh_table()
    def go_next():
        if state["page"]<state["total_pages"]:
            state["page"]+=1; refresh_table()
    btn_prev.config(command=go_prev); btn_next.config(command=go_next)

    def export_pdf_with_notice():
        if messagebox.askokcancel("Export PDF", "Le PDF n'inclura pas la colonne Type. Continuer ?"):
            do_export("pdf", date_from_var.get(), date_to_var.get(), parent, movement_type_var.get())

    btn_search.config(command=lambda: load_data_and_refresh(page=1))
    btn_refresh.config(command=lambda: load_data_and_refresh(page=state.get("page", 1)))
    btn_export_xls.config(command=lambda: do_export("xlsx", date_from_var.get(), date_to_var.get(), parent, movement_type_var.get()))
    btn_export_pdf.config(command=export_pdf_with_notice)

    load_data_and_refresh(page=1)

def do_export(kind: str, date_from_value: str, date_to_value: str, parent, movement_type_filter: str = "ALL"):
    from_d = parse_date_input(date_from_value)
    to_d = parse_date_input(date_to_value)
    data = query_declared_articles(date_from=from_d, date_to=to_d, movement_type_filter=movement_type_filter)
    if not data:
        messagebox.showinfo("Aucun résultat","Aucun enregistrement à exporter pour les filtres choisis.", parent=parent)
        return
    if kind == "xlsx":
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx"),("CSV","*.csv")], title="Exporter Excel", parent=parent)
        if not path: return
        ok, err = export_to_excel_pandas(data, path)
        if ok: messagebox.showinfo("Export", f"Export Excel réussi: {os.path.basename(path)}", parent=parent)
        else: messagebox.showerror("Export échoué", f"Erreur: {err}", parent=parent)
    else:
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")], title="Exporter PDF", parent=parent)
        if not path: return
        ok, err = export_to_pdf_reportlab(data, path)
        if ok: messagebox.showinfo("Export", f"PDF créé: {os.path.basename(path)}", parent=parent)
        else: messagebox.showerror("Export échoué", f"Erreur: {err}", parent=parent)

def main():
    root = tk.Tk()
    root.title("Articles déclarés (OBR=1) - Viewer (compact)")
    root.geometry("1280x820")
    show_obr_articles(root)
    root.mainloop()

if __name__ == "__main__":
    main()
