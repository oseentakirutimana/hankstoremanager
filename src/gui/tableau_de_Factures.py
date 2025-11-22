# aff# afficher_liste_factures.py
# -*- coding: utf-8 -*-
"""
Module complet : afficher_liste_factures + helpers nécessaires

Mises à jour demandées :
- Export PDF: retirer invoice_identifier, ajouter invoice_number, invoice_date, invoice_type
- Export PDF: inclure les colonnes d’articles (item_designation, quantity, unit_price_used, unit_price_nvat, vat_amount, line_total_amount)
- Export Excel: même logique, avec onglets séparés "Factures" et "Articles" pour une structure propre
- PDF: éviter le chevauchement des colonnes (colWidths adaptés, alignements à droite pour les montants)
- Pagination par défaut : 20 éléments par page (configurable via variable d'environnement PAGE_SIZE)

Modifications supplémentaires appliquées :
- PVT HTVA (Partie HT) calculé explicitement comme somme(qty * unit_price_used) partout dans les modales imprimables (printable, A4, mobile, EBMS modal).
- CheckboxFlowable (Oui / Non) détermine maintenant l'état coché à partir des champs vat_taxpayer / vat_customer_payer (accepte 1, "1", True, "True").
"""
import os
import math
import logging
import sqlite3
import requests
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from typing import Optional, List, Dict, Tuple
import unicodedata
import hashlib
import sys
import subprocess
import tempfile
from typing import Optional

# reportlab platypus/table helpers used by generate_invoice_pdf
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import Flowable
    from reportlab.lib import colors
    from reportlab.lib.units import mm
except Exception:
    # reportlab optional; functions will check for availability
    SimpleDocTemplate = None
    Paragraph = Spacer = Table = TableStyle = PageBreak = Flowable = None
    getSampleStyleSheet = ParagraphStyle = None
    A4 = landscape = colors = mm = None

# fonctions demandées par l’utilisateur (assure-toi qu’elles existent)
from api.obr_client import obtenir_token_auto, get_system_id, checkTIN
from utils.obr_db_helpers import get_next_invoice_number
from database.connection import get_connection

# Optionnel
try:
    from tkcalendar import DateEntry
except Exception:
    DateEntry = None
try:
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
except Exception:
    pd = None

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# -----------------------
# Configuration et constantes UI
# -----------------------
CONTENT_BG = "#f6f8fa"
CARD_BG = "#ffffff"
CONTOUR_BG = "#e6eef9"
LABEL_FG = "#1f2937"
TITLE_FG = "#0b3d91"
FONT_TITLE = ("Segoe UI", 16, "bold")
FONT_LABEL = ("Segoe UI", 11)
FONT_CELL = ("Segoe UI", 10)
ROW_ALT = "#fbfdff"

# Table visual constants
TABLE_HEADER_BG = "#eef6ff"
TABLE_CELL_BG = "white"
TABLE_HEADER_FONT = ("Segoe UI", 10, "bold")
TABLE_CELL_FONT = ("Segoe UI", 10)
TABLE_CELL_PADX = 6
TABLE_CELL_PADY = 6

# DEFAULT PAGE SIZE: can be overridden with the PAGE_SIZE env var
try:
    PAGE_SIZE = int(os.environ.get("PAGE_SIZE", "20"))  # default 20
except Exception:
    PAGE_SIZE = 20
PAGE_SIZE = max(1, min(PAGE_SIZE, 500))

# Colonnes affichées dans la liste (UI)
INVOICE_COLUMNS = [
    ("invoice_date", "Date", 18),
    ("invoice_number", "Numéro", 16),
    ("customer_name", "Client", 26),
    ("facture_statut", "Statut", 12),
    ("invoice_type", "Type", 16),
]

# -----------------------
# Styles ttk colorés
# -----------------------
# Appliquer et forcer les couleurs sur les styles TButton
_style = ttk.Style()

# Tenter d'utiliser un thème qui respecte les couleurs (clam/alt/default) sinon fallback
for preferred in ("clam", "alt", "default"):
    try:
        _style.theme_use(preferred)
        break
    except Exception:
        pass

# Helper for safe configure + map
def _cfg(name, **kwargs):
    try:
        _style.configure(name, **kwargs)
    except Exception:
        pass

def _map(name, option, values):
    try:
        _style.map(name, **{option: values})
    except Exception:
        pass

# Redéfinir layout simple pour s'assurer que background est pris en compte
try:
    btn_layout = [('Button.border', {'children':
                    [('Button.padding', {'children':
                        [('Button.label', {'sticky': 'nswe'})],
                    'sticky': 'nswe'})],
                'sticky': 'nswe'})]
    _style.layout('Primary.TButton', btn_layout)
    _style.layout('Success.TButton', btn_layout)
    _style.layout('Warning.TButton', btn_layout)
    _style.layout('Danger.TButton', btn_layout)
    _style.layout('Default.TButton', btn_layout)
except Exception:
    pass

# Configure styles (foreground et background)
_cfg("Primary.TButton", foreground="white", background="#2563eb", padding=6, relief="flat")
_map("Primary.TButton", "background", [("active", "#1e40af"), ("disabled", "#9dbfe8")])
_map("Primary.TButton", "foreground", [("disabled", "#f0f0f0")])

_cfg("Success.TButton", foreground="white", background="#16a34a", padding=6, relief="flat")
_map("Success.TButton", "background", [("active", "#15803d"), ("disabled", "#9ccfa0")])

_cfg("Warning.TButton", foreground="black", background="#f59e0b", padding=6, relief="flat")
_map("Warning.TButton", "background", [("active", "#d97706"), ("disabled", "#f3c37a")])

_cfg("Danger.TButton", foreground="white", background="#dc3545", padding=6, relief="flat")
_map("Danger.TButton", "background", [("active", "#b91c1c"), ("disabled", "#e79a9f")])

_cfg("Default.TButton", foreground="black", background="#e5e7eb", padding=6, relief="flat")
_map("Default.TButton", "background", [("active", "#d1d5db"), ("disabled", "#f0f0f0")])

# Make sure ttk.Button uses the Primary style by default when no style is passed (optional)
try:
    _style.configure("TButton", padding=6, relief="flat")
except Exception:
    pass


# -----------------------
# Utilitaires UI
# -----------------------
def _center_window(win: tk.Toplevel, parent: Optional[tk.Widget], w: int, h: int):
    win.update_idletasks()
    if parent:
        try:
            px = parent.winfo_rootx(); py = parent.winfo_rooty()
            pw = parent.winfo_width(); ph = parent.winfo_height()
            x = px + max(0, (pw - w) // 2)
            y = py + max(0, (ph - h) // 2)
        except Exception:
            sw = win.winfo_screenwidth(); sh = win.winfo_screenheight()
            x = (sw - w) // 2; y = (sh - h) // 2
    else:
        sw = win.winfo_screenwidth(); sh = win.winfo_screenheight()
        x = (sw - w) // 2; y = (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")

def _modal_buttons(parent_frame: tk.Frame, buttons: List[Tuple[str, callable, str]], align: str = "right"):
    frame = tk.Frame(parent_frame, bg=parent_frame.cget("bg"))
    frame.pack(fill="x", padx=8, pady=8)
    if align == "right":
        spacer = tk.Frame(frame, bg=parent_frame.cget("bg")); spacer.pack(side="left", expand=True)
        for text, cmd, style in buttons:
            style_name = style if style.endswith(".TButton") else f"{style}.TButton"
            b = ttk.Button(frame, text=text, command=cmd, style=style_name)
            b.pack(side="right", padx=6)
    else:
        for text, cmd, style in buttons:
            style_name = style if style.endswith(".TButton") else f"{style}.TButton"
            b = ttk.Button(frame, text=text, command=cmd, style=style_name)
            b.pack(side="left", padx=6)


def _open_file_with_default_app(path: str) -> None:
    """
    Ouvre le fichier avec l'application par défaut selon la plateforme.
    - Windows : os.startfile(path)
    - macOS   : open path
    - Linux   : xdg-open ou gio open
    Lève une exception si l'ouverture échoue.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    if sys.platform.startswith("win"):
        # may raise OSError on failure
        os.startfile(path)
        return

    if sys.platform == "darwin":
        proc = subprocess.Popen(["open", path])
        proc.poll()
        return

    # Generic Unix-like
    for cmd in (["xdg-open", path], ["gio", "open", path]):
        try:
            proc = subprocess.Popen(cmd)
            proc.poll()
            return
        except FileNotFoundError:
            continue
    # If we reach here, no opener found
    raise RuntimeError("Aucune application d'ouverture trouvée (xdg-open/gio absent)")

def _print_file_direct(path: str) -> bool:
    """
    Tente d'envoyer directement le PDF à l'imprimante par défaut.
    Retourne True si une commande a été lancée sans erreur immédiate.
    Comportement :
      - Windows : os.startfile(path, "print") (peut ouvrir le viewer)
      - macOS   : lp path
      - Linux   : lpr path ou lp path
    Note : La réussite dépend de la configuration système et des permissions.
    """
    if not os.path.exists(path):
        return False

    try:
        if sys.platform.startswith("win"):
            try:
                os.startfile(path, "print")
                return True
            except Exception:
                return False

        if sys.platform == "darwin":
            subprocess.check_call(["lp", path])
            return True

        # Linux / other Unix
        for cmd in (["lpr", path], ["lp", path]):
            try:
                subprocess.check_call(cmd)
                return True
            except (FileNotFoundError, subprocess.CalledProcessError):
                continue
        return False
    except Exception:
        return False


# -----------------------
# Dates / format
# -----------------------
def parse_date_input(s: Optional[str]) -> Optional[str]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None

def format_date_short(txt: Optional[str]) -> str:
    if not txt:
        return ""
    try:
        dt = datetime.strptime(txt, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(txt)

# -----------------------
# Signature helpers
# -----------------------
def _sha1_hex_normalized(text: str) -> str:
    # OBR expects SHA-256 of the normalized invoice signature (document indicates SHA256)
    try:
        s = unicodedata.normalize("NFKD", (text or "").strip()).encode("utf-8")
        return hashlib.sha256(s).hexdigest()
    except Exception:
        return ""

def _build_obr_invoice_signature(tp: dict, invoice_number: str, sig_dt_ui: str = None):
    now_dt = datetime.now()
    try:
        if sig_dt_ui:
            dt_parsed = datetime.strptime(sig_dt_ui.strip(), "%Y-%m-%d %H:%M:%S")
            sig_date_field = dt_parsed.strftime("%Y-%m-%d %H:%M:%S")
            timestamp_compact = dt_parsed.strftime("%Y%m%d%H%M%S")
        else:
            sig_date_field = now_dt.strftime("%Y-%m-%d %H:%M:%S")
            timestamp_compact = now_dt.strftime("%Y%m%d%H%M%S")
    except Exception:
        sig_date_field = now_dt.strftime("%Y-%m-%d %H:%M:%S")
        timestamp_compact = now_dt.strftime("%Y%m%d%H%M%S")
    nif = (tp.get("tp_TIN") or "").strip() or "UNKNOWN"
    system_id = str(get_system_id() or "").strip() or "UNKNOWN"
    inv_num = (invoice_number or "").strip() or get_next_invoice_number()
    invoice_signature = f"{nif}/{system_id}/{timestamp_compact}/{inv_num}"
    electronic_signature = _sha1_hex_normalized(invoice_signature)
    return invoice_signature, sig_date_field, electronic_signature

# -----------------------
# PDF small helpers
# -----------------------
def format_date_short_pdf(txt: Optional[str]) -> str:
    try:
        if not txt:
            return ""
        dt = datetime.strptime(txt, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(txt or "")

def format_money_pdf(v) -> str:
    try:
        return f"{float(v):,.2f}"
    except Exception:
        return "0.00"

# -----------------------
# generate_invoice_pdf (A4, header + table + footer)
# -----------------------
def format_money(v) -> str:
    try:
        return f"{float(v):,.2f}"
    except Exception:
        return str(v or "")


# Helper utilitaire pour normaliser l'état assujetti (placer avec les autres helpers PDF)
def _is_assujetti(val) -> bool:
    """
    Interprète de façon défensive les valeurs possibles pour 'assujetti à la TVA'.
    Accepte : True/False, 1/"1", "true"/"True"/"TRUE", "yes"/"Yes", "y"/"Y".
    """
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    try:
        s = str(val).strip()
    except Exception:
        return False
    return s in ("1", "true", "True", "TRUE", "yes", "Yes", "Y", "y")


# Remplacement complet de la classe CheckboxFlowable
class CheckboxFlowable(Flowable if Flowable is not None else object):
    """
    Flowable that draws two small graphic checkboxes with labels "Oui" and "Non",
    placing an X inside the checked one.
    Normalise strictement le flag checked en bool pour éviter les surprises.
    """
    def __init__(self, checked: bool, box_size: float = 9, gap: float = 8, label_fontsize: int = 9):
        # normalization to strict boolean
        try:
            self.checked = bool(checked)
        except Exception:
            self.checked = False
        self.box_size = float(box_size)
        self.gap = float(gap)
        try:
            self.label_fontsize = int(label_fontsize)
        except Exception:
            self.label_fontsize = 9

        # conservative width estimate: two boxes + label space + gap
        label_w = max(18, self.label_fontsize * 2)
        per_block = self.box_size + 4 + label_w
        self.width = per_block * 2 + self.gap
        self.height = max(self.box_size, self.label_fontsize) + 2

    def wrap(self, availWidth, availHeight):
        return min(self.width, availWidth), min(self.height, availHeight)

    def draw(self):
        try:
            c = self.canv
            c.saveState()
            x = 0
            y_top = self.height
            # first box (Oui)
            c.rect(x, y_top - self.box_size, self.box_size, self.box_size, stroke=1, fill=0)
            if self.checked:
                inset = max(1, self.box_size * 0.18)
                c.setLineWidth(max(0.6, self.box_size * 0.12))
                c.line(x + inset, y_top - self.box_size + inset, x + self.box_size - inset, y_top - inset)
                c.line(x + inset, y_top - inset, x + self.box_size - inset, y_top - self.box_size + inset)
                c.setLineWidth(0.5)
            c.setFont("Helvetica", self.label_fontsize)
            label_x = x + self.box_size + 4
            label_y = y_top - (self.box_size * 0.32)
            c.drawString(label_x, label_y, "Oui")

            # second box (Non)
            second_x = label_x + max((self.label_fontsize * 5), 24) + self.gap
            c.rect(second_x, y_top - self.box_size, self.box_size, self.box_size, stroke=1, fill=0)
            if not self.checked:
                inset = max(1, self.box_size * 0.18)
                c.setLineWidth(max(0.6, self.box_size * 0.12))
                c.line(second_x + inset, y_top - self.box_size + inset, second_x + self.box_size - inset, y_top - inset)
                c.line(second_x + inset, y_top - inset, second_x + self.box_size - inset, y_top - self.box_size + inset)
                c.setLineWidth(0.5)
            c.drawString(second_x + self.box_size + 4, label_y, "Non")
        except Exception:
            try:
                c.restoreState()
            except Exception:
                pass
        finally:
            try:
                c.restoreState()
            except Exception:
                pass


# Remplacement complet de generate_invoice_pdf (A4)
def generate_invoice_pdf(data: Dict, filename: str) -> Tuple[bool, Optional[str]]:
    if SimpleDocTemplate is None:
        return False, "reportlab non installé"
    try:
        styles = getSampleStyleSheet()
        normal = styles["Normal"]
        title_style = ParagraphStyle("InvTitle", parent=styles["Heading1"], alignment=1, fontSize=14, spaceAfter=6)
        small = ParagraphStyle("Small", parent=normal, fontSize=9, leading=11)
        item_para_style = ParagraphStyle("ItemPara", parent=normal, fontSize=9, leading=11)

        doc = SimpleDocTemplate(filename, pagesize=A4,
                                leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
        elems: List = []

        inv_num = data.get("invoice_number") or "........"
        inv_date = data.get("invoice_date") or ""
        title_text = f"Facture n° {inv_num}   du {inv_date}"
        elems.append(Paragraph(title_text, title_style))
        elems.append(Spacer(1, 6))

        tp = (data.get("tp") or {}) or {}
        client = (data.get("client") or {}) or {}

        tp_is_assujetti = _is_assujetti(tp.get("vat_taxpayer") or tp.get("tp_vat_taxpayer"))
        client_is_assujetti = _is_assujetti(client.get("vat_customer_payer") or client.get("vat_customer"))

        # Vendeur
        left_para_data = Paragraph("<br/>".join([
            "<b>A. Identification du vendeur</b>",
            f"Nom et prénom ou Raison sociale : {tp.get('tp_name','')}",
            f"Centre fiscal : {tp.get('tp_fiscal_center','')}",
            f"NIF : {tp.get('tp_TIN','')}",
            f"Secteur d'activités : {tp.get('tp_activity_sector','')}",
        ]), small)

        tva_row_table = Table(
            [[Paragraph("Assujetti à la TVA :", small), CheckboxFlowable(bool(tp_is_assujetti))]],
            colWidths=[doc.width * 0.27, doc.width * 0.28]
        )
        tva_row_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 2),
            ("LEFTPADDING", (1, 0), (1, 0), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))

        left_cell = Table([[left_para_data], [tva_row_table]], colWidths=[doc.width * 0.55])
        left_cell.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        right_para = Paragraph("<br/>".join([
            f"Registre de Commerce N°: {tp.get('tp_trade_number','')}",
            f"Forme juridique : {tp.get('tp_legal_form','')}",
            f"B.P : {tp.get('tp_postal_number','')} , Tél : {tp.get('tp_phone_number','')}",
            f"Av. : {tp.get('tp_address_avenue','')} , Rue : {tp.get('tp_address_rue','')} , N° {tp.get('tp_address_number','')}",
            f"Commune : {tp.get('tp_address_commune','')} , Quartier : {tp.get('tp_address_quartier','')}"
        ]), small)

        vendor_tbl = Table([[left_cell, right_para]], colWidths=[doc.width * 0.5, doc.width * 0.5])
        vendor_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        elems.append(vendor_tbl)
        elems.append(Spacer(1, 6))

        # Client
        client_para_data = Paragraph("<br/>".join([
            "<b>B. Le client</b>",
            f"Nom et prénom ou Raison sociale: {client.get('customer_name','')}",
            f"NIF : {client.get('customer_TIN','')}",
            f"Résident à : {client.get('customer_address','')}",
        ]), small)

        client_tva_row_table = Table(
            [[Paragraph("Assujetti à la TVA :", small), CheckboxFlowable(bool(client_is_assujetti))]],
            colWidths=[doc.width * 0.27, doc.width * 0.42]
        )
        client_tva_row_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 2),
            ("LEFTPADDING", (1, 0), (1, 0), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))

        client_cell = Table([[client_para_data], [client_tva_row_table]], colWidths=[doc.width])
        client_cell.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elems.append(client_cell)
        elems.append(Spacer(1, 8))

        elems.append(Paragraph("doit pour ce qui suit :", normal))
        elems.append(Spacer(1, 6))

        # Items table
        items = data.get("invoice_items", []) or []
        header = ["Nature de l'article ou service*", "Qté*", "PU*", "PVHTVA"]
        table_data: List[List] = [header]
        for it in items:
            desc = Paragraph(str(it.get("item_designation", "") or ""), item_para_style)
            qty = str(it.get("item_quantity", "") or "")
            pu = format_money(it.get("item_unit_price", "") or 0)
            total = format_money(it.get("item_total_amount", "") or 0)
            table_data.append([desc, qty, pu, total])

        if len(table_data) == 1:
            table_data.append(["", "", "", ""])

        col_widths = [doc.width * 0.58, doc.width * 0.12, doc.width * 0.15, doc.width * 0.15]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#263238")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f0ff")),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        elems.append(tbl)
        elems.append(Spacer(1, 8))

        # Totals (defensive)
        totals = data.get("totals") or {}
        subtotal = totals.get("subtotal", None)
        vat = totals.get("vat", None)
        grand_total = totals.get("grand_total", None)

        if subtotal in (None, ""):
            s_sub = 0.0
            s_vat = 0.0
            for it in items:
                try:
                    qty = float(it.get("item_quantity") or it.get("quantity") or 0)
                except Exception:
                    try: qty = float(str(it.get("item_quantity") or it.get("quantity") or 0).replace(",", "."))
                    except Exception: qty = 0.0
                try:
                    pu = float(it.get("item_unit_price") or it.get("unit_price_used") or it.get("item_unit_price") or it.get("item_price") or 0)
                except Exception:
                    try: pu = float(str(it.get("item_unit_price") or it.get("unit_price_used") or it.get("item_price") or 0).replace(",", "."))
                    except Exception: pu = 0.0
                line_ht = qty * pu
                try:
                    vat_i = float(it.get("item_vat_amount") or it.get("vat") or it.get("vat_amount") or 0)
                except Exception:
                    try: vat_i = float(str(it.get("item_vat_amount") or it.get("vat") or 0).replace(",", "."))
                    except Exception: vat_i = 0.0
                s_sub += line_ht
                s_vat += vat_i
            subtotal = round(s_sub, 2)
            vat = round(s_vat, 2) if vat in (None, "") else vat
            grand_total = round(subtotal + (vat or 0.0), 2) if grand_total in (None, "") else grand_total

        totals_table_data = [
            ["PVT HTVA:", format_money(subtotal)],
            ["TVA:", format_money(vat or 0)],
            ["Total TVAC:", format_money(grand_total or 0)]
        ]
        tot_tbl = Table(totals_table_data, colWidths=[doc.width * 0.25, doc.width * 0.75])
        tot_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bfc9d9")),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(tot_tbl)
        elems.append(Spacer(1, 6))

        sig = data.get("electronic_signature") or ""
        if sig:
            elems.append(Paragraph(f"signature : {sig}", small))
            elems.append(Spacer(1, 6))

        footer_notes = data.get("footer_notes", [
            "* Mention obligatoire",
            "N.B: Les non assujettis à la TVA ne remplissent pas les deux dernières lignes"
        ]) or []
        for note in footer_notes:
            elems.append(Paragraph(note, small))

        doc.build(elems)
        return True, None

    except Exception as e:
        try:
            logging.getLogger(__name__).exception("Generate PDF failed: %s", e)
        except Exception:
            pass
        return False, str(e)


# Remplacement complet de generate_invoice_mobile_pdf (mobile)
def generate_invoice_mobile_pdf(data: Dict, filename: str) -> Tuple[bool, Optional[str]]:
    """
    Générateur compact optimisé pour imprimantes mobiles (~80 mm).
    Lit data (payload construit par _build_local_payload_from_minimal ou EBMS)
    et produit un PDF mobile. Utilise _is_assujetti et CheckboxFlowable définis
    dans le module.
    """
    if SimpleDocTemplate is None:
        return False, "reportlab non installé"

    try:
        page_w, page_h = (80 * mm, 300 * mm)
        doc = SimpleDocTemplate(
            filename,
            pagesize=(page_w, page_h),
            leftMargin=4 * mm, rightMargin=4 * mm, topMargin=4 * mm, bottomMargin=4 * mm
        )

        styles = getSampleStyleSheet()
        normal = styles["Normal"]
        title_style = ParagraphStyle("MobTitle", parent=styles["Heading2"], alignment=1, fontSize=10, spaceAfter=4)
        small = ParagraphStyle("MobSmall", parent=normal, fontSize=7.5, leading=9)
        item_para_style = ParagraphStyle("MobItemPara", parent=normal, fontSize=7.5, leading=8, spaceAfter=0)
        footer_style = ParagraphStyle("MobFooter", parent=small, fontSize=6.5, leading=7)

        elems: List = []

        inv_num = data.get("invoice_number") or ""
        inv_date = data.get("invoice_date") or ""
        title_text = f"Facture n° {inv_num}   du {format_date_short_pdf(inv_date) if inv_date else ''}"
        elems.append(Paragraph(title_text, title_style))
        elems.append(Spacer(1, 4))

        tp = (data.get("tp") or {}) or {}
        client = (data.get("client") or {}) or {}

        # Defensive normalization of assujetti flags (use module-level _is_assujetti)
        tp_val = tp.get("vat_taxpayer") if tp.get("vat_taxpayer") not in (None, "") else tp.get("tp_vat_taxpayer")
        client_val = client.get("vat_customer_payer") if client.get("vat_customer_payer") not in (None, "") else (client.get("vat_customer") or client.get("vat_customer_payer"))

        tp_is_assujetti = _is_assujetti(tp_val)
        client_is_assujetti = _is_assujetti(client_val)

        left_para = Paragraph("<br/>".join([
            "<b>A. Identification du vendeur</b>",
            f"Nom et prénom ou Raison sociale : {tp.get('tp_name','')}",
            f"B.P : {tp.get('tp_postal_number','')} , Tél : {tp.get('tp_phone_number','')}",
            f"Registre de Commerce N°: {tp.get('tp_trade_number','')}",
            f"Secteur d'activités : {tp.get('tp_activity_sector','')}",
        ]), small)

        usable = page_w - 8 * mm
        tva_row_table = Table(
            [[Paragraph("Assujetti à la TVA :", small), CheckboxFlowable(tp_is_assujetti, box_size=7, gap=6, label_fontsize=7)]],
            colWidths=[usable * 0.45, usable * 0.35]
        )
        tva_row_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        left_cell = Table([[left_para], [tva_row_table]], colWidths=[usable * 0.6])
        left_cell.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        right_para = Paragraph("<br/>".join([
            f"NIF : {tp.get('tp_TIN','')}",
            f"Forme juridique : {tp.get('tp_legal_form','')}",
            f"Centre fiscal : {tp.get('tp_fiscal_center','')}",
            f"Commune : {tp.get('tp_address_commune','')} , Quartier : {tp.get('tp_address_quartier','')}"
        ]), small)

        vendor_tbl = Table([[left_cell, right_para]], colWidths=[usable * 0.6, usable * 0.4])
        vendor_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        elems.append(vendor_tbl)
        elems.append(Spacer(1, 4))

        client_para = Paragraph("<br/>".join([
            "<b>B. Le client</b>",
            f"Nom et prénom ou Raison sociale: {client.get('customer_name','')}",
            f"NIF : {client.get('customer_TIN','')}",
            f"Résident à : {client.get('customer_address','')}",
        ]), small)

        client_tva_row_table = Table(
            [[Paragraph("Assujetti à la TVA :", small), CheckboxFlowable(client_is_assujetti, box_size=7, gap=6, label_fontsize=7)]],
            colWidths=[usable * 0.45, usable * 0.35]
        )
        client_tva_row_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        client_cell = Table([[client_para], [client_tva_row_table]], colWidths=[usable])
        client_cell.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elems.append(client_cell)
        elems.append(Spacer(1, 6))

        elems.append(Paragraph("doit pour ce qui suit :", small))
        elems.append(Spacer(1, 4))

        # Items table
        items = data.get("invoice_items", []) or []
        header = ["Désignation", "Qté", "PU", "Total"]
        table_data: List[List] = [header]
        for it in items:
            desc_para = Paragraph(str(it.get("item_designation") or it.get("designation") or ""), item_para_style)
            qty = str(it.get("item_quantity") or it.get("quantity") or "")
            pu = format_money_pdf(it.get("item_unit_price") or it.get("item_price") or 0)
            total = format_money_pdf(it.get("item_total_amount") or it.get("line_total_amount") or 0)
            table_data.append([desc_para, Paragraph(qty, item_para_style), Paragraph(pu, item_para_style), Paragraph(total, item_para_style)])

        if len(table_data) == 1:
            table_data.append(["", "", "", ""])

        col_widths = [usable * 0.46, usable * 0.16, usable * 0.19, usable * 0.19]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#333333")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f4ff")),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, 0), 2),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ]))
        elems.append(tbl)
        elems.append(Spacer(1, 6))

        # Totals computation (defensive)
        totals = data.get("totals") or {}
        subtotal = totals.get("subtotal", None)
        vat = totals.get("vat", None)
        grand_total = totals.get("grand_total", None)

        if subtotal is None or vat is None or grand_total is None:
            s_sub, s_vat = 0.0, 0.0
            for it in items:
                def _to_float_safe(x):
                    try:
                        if x is None or x == "":
                            return 0.0
                        if isinstance(x, (int, float)):
                            return float(x)
                        t = str(x).strip().replace(" ", "").replace("\u00A0", "")
                        if t.count(",") > 0 and t.count(".") == 0:
                            t = t.replace(",", ".")
                        else:
                            t = t.replace(",", "")
                        return float(t)
                    except Exception:
                        return 0.0
                qty = _to_float_safe(it.get("item_quantity") or it.get("quantity") or 0)
                pu_val = _to_float_safe(it.get("item_unit_price") or it.get("unit_price_used") or it.get("item_price") or 0)
                line_ht = qty * pu_val
                vat_amt = _to_float_safe(it.get("item_vat_amount") or it.get("vat") or it.get("vat_amount") or 0)
                s_sub += line_ht
                s_vat += vat_amt
            subtotal = s_sub
            vat = s_vat
            grand_total = subtotal + vat

        totals_table_data = [
            [Paragraph("PVT HTVA:", small), Paragraph(format_money_pdf(subtotal), small)],
            [Paragraph("TVA:", small), Paragraph(format_money_pdf(vat), small)],
            [Paragraph("<b>Total TVAC:</b>", small), Paragraph(f"<b>{format_money_pdf(grand_total)}</b>", small)]
        ]
        tot_tbl = Table(totals_table_data, colWidths=[usable * 0.20, usable * 0.70], hAlign="RIGHT")
        tot_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#bfc9d9")),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elems.append(tot_tbl)
        elems.append(Spacer(1, 4))

        sig = data.get("electronic_signature") or ""
        if not sig:
            inv_ident = data.get("invoice_identifier") or ""
            if inv_ident:
                sig = inv_ident
        if sig:
            elems.append(Paragraph(f"signature : {sig}", small))
            elems.append(Spacer(1, 4))

        footer_notes = data.get("footer_notes", [
            "* Mention obligatoire",
            "N.B: Les non assujettis à la TVA ne remplissent pas les deux dernières lignes"
        ])
        for note in footer_notes:
            elems.append(Paragraph(str(note), footer_style))

        doc.build(elems)
        return True, None

    except Exception as e:
        logger.exception("Generate mobile PDF failed: %s", e)
        return False, str(e)


# -----------------------
# Modal imprimable compact (petites imprimantes)
# -----------------------
def _show_invoices_modal_printable(invoices: list, parent=None, source_is_ebms: bool = True):
    """
    Affiche une modal imprimable pour une ou plusieurs factures.

    - invoices: liste d'objets invoice (forme EBMS ou payload minimal)
    - source_is_ebms: si True on utilise les champs fournis; si False on enrichit depuis la DB
    Préconditions (doivent exister dans le module) :
      - tkinter : tk, ttk, messagebox, filedialog
      - utilitaires : _center_window, format_date_short_pdf, format_money_pdf
      - PDF helpers : generate_invoice_mobile_pdf, generate_invoice_pdf, _generate_pdf_from_local
      - DB helper : _build_local_payload_from_minimal
      - OS helpers : tempfile, os, _open_file_with_default_app, _print_file_direct
    """
    popup = tk.Toplevel(parent)
    popup.title("Détails facture - Impression")
    popup.transient(parent)
    popup.grab_set()
    width, height = 720, 420
    _center_window(popup, parent, width, height)

    frame_root = tk.Frame(popup, bg="white", padx=8, pady=8)
    frame_root.pack(fill="both", expand=True)

    def _fmt_num(x):
        try:
            if x is None or x == "":
                return ""
            return f"{float(x):.2f}"
        except Exception:
            s = str(x).replace(",", ".").replace(" ", "")
            try:
                return f"{float(s):.2f}"
            except Exception:
                return str(x)

    for inv in invoices:
        panel = tk.Frame(frame_root, bg="white", bd=0, relief="flat", padx=6, pady=6)
        panel.pack(fill="both", expand=False, pady=(4, 8))

        # Build or enrich payload depending on source flag
        if source_is_ebms:
            payload = {
                "invoice_number": inv.get("invoice_registered_number") or inv.get("invoice_number") or inv.get("invoice_identifier", ""),
                "invoice_date": inv.get("invoice_registered_date") or inv.get("invoice_date", ""),
                "invoice_identifier": inv.get("invoice_identifier") or "",
                "tp": {
                    "tp_name": inv.get("tp_name") or inv.get("seller_name") or "",
                    "tp_TIN": inv.get("tp_TIN") or "",
                    "tp_trade_number": inv.get("tp_trade_number") or "",
                    "tp_phone_number": inv.get("tp_phone_number") or "",
                    "tp_postal_number": inv.get("tp_postal_number") or "",
                    "tp_address_commune": inv.get("tp_address_commune") or "",
                    "tp_address_quartier": inv.get("tp_address_quartier") or "",
                    "tp_activity_sector": inv.get("tp_activity_sector") or "",
                    "vat_taxpayer": inv.get("vat_taxpayer") or inv.get("tp_vat_taxpayer") or 0
                },
                "client": {
                    "customer_name": inv.get("buyer_name") or inv.get("customer_name") or inv.get("client_name") or "",
                    "customer_TIN": inv.get("buyer_tin") or inv.get("customer_TIN") or "",
                    "customer_address": inv.get("buyer_address") or inv.get("customer_address") or "",
                    "vat_customer_payer": inv.get("vat_customer_payer") or 0
                },
                "invoice_items": inv.get("invoice_items") or inv.get("items") or [],
                "totals": inv.get("totals") or {},
                "electronic_signature": inv.get("invoice_identifier") or inv.get("signature") or ""
            }
        else:
            payload = _build_local_payload_from_minimal(inv)

        # Header (single line: number + date)
        title = payload.get("invoice_number", "")
        date_txt = payload.get("invoice_date", "")
        header_text = f"Facture: {title}"
        if date_txt:
            header_text += f"   du {format_date_short_pdf(date_txt)}"
        tk.Label(panel, text=header_text, font=("Segoe UI", 11, "bold"), bg="white").pack(anchor="w")

        # Client and optional status
        cust_name = payload.get("client", {}).get("customer_name", "")
        if cust_name:
            tk.Label(panel, text=f"Client: {cust_name}", bg="white").pack(anchor="w")
        statut = inv.get("facture_statut") or ""
        if statut:
            tk.Label(panel, text=f"Statut: {statut}", bg="white").pack(anchor="w")
        tk.Label(panel, text="", bg="white").pack(anchor="w", pady=(0, 4))  # small spacer

        # Items table header
        tbl = tk.Frame(panel, bg="white")
        tbl.pack(fill="both", expand=True)
        hdrs = ["Désignation", "Qté", "PU", "Total ligne"]
        for ci, h in enumerate(hdrs):
            tk.Label(tbl, text=h, bg="#eef6ff", font=("Segoe UI", 9, "bold"),
                     bd=1, relief="solid", padx=6, pady=6).grid(row=0, column=ci, sticky="nsew")

        items = payload.get("invoice_items") or []
        for ri, it in enumerate(items, start=1):
            des = it.get("item_designation") or it.get("designation") or ""
            qty = it.get("item_quantity") or it.get("quantity") or ""
            # prefer unit_price_used / item_unit_price / item_price
            pu = it.get("item_unit_price") or it.get("unit_price_used") or it.get("unit_price_nvat") or it.get("item_price") or ""
            total = it.get("item_total_amount") or it.get("line_total_amount") or it.get("item_total") or ""
            vals = [des, _fmt_num(qty), _fmt_num(pu), _fmt_num(total)]
            for ci, v in enumerate(vals):
                anchor = "e" if ci >= 1 else "w"
                tk.Label(tbl, text=str(v), bg="white", anchor=anchor,
                         bd=1, relief="solid", padx=6, pady=6).grid(row=ri, column=ci, sticky="nsew")

        for c in range(len(hdrs)):
            tbl.grid_columnconfigure(c, weight=1)

        # Totals: prefer payload['totals'], else compute defensively
        totals = payload.get("totals") or {}
        subtotal = totals.get("subtotal", None)
        vat = totals.get("vat", None)
        grand = totals.get("grand_total", None)

        if subtotal in (None, ""):
            s_sub = 0.0
            s_vat = 0.0
            for it in items:
                try:
                    qty = float(it.get("item_quantity") or it.get("quantity") or 0)
                except Exception:
                    try:
                        qty = float(it.get("item_quantity") or it.get("quantity") or 0)
                    except Exception:
                        qty = 0.0
                try:
                    pvu = float(it.get("item_unit_price") or it.get("unit_price_used") or it.get("unit_price_nvat") or it.get("item_price") or 0)
                except Exception:
                    try:
                        pvu = float(str(it.get("item_unit_price") or it.get("unit_price_used") or it.get("item_price") or 0).replace(",", "."))
                    except Exception:
                        pvu = 0.0
                ln = qty * pvu
                try:
                    vat_i = float(it.get("item_vat_amount") or it.get("vat") or 0)
                except Exception:
                    try:
                        vat_i = float(str(it.get("item_vat_amount") or it.get("vat")).replace(",", "."))
                    except Exception:
                        vat_i = 0.0
                s_sub += ln
                s_vat += vat_i
            subtotal = s_sub
            vat = s_vat
            grand = subtotal + vat

        totals_frame = tk.Frame(panel, bg="white")
        totals_frame.pack(fill="x", pady=(6, 0))
        tk.Label(totals_frame, text=f"PVT HTVA: {subtotal:.2f}", bg="white", anchor="e").pack(anchor="w")
        tk.Label(totals_frame, text=f"TVA: {vat:.2f}", bg="white", anchor="e").pack(anchor="w")
        tk.Label(totals_frame, text=f"Total TVAC: {grand:.2f}", bg="white",
                 font=("Segoe UI", 9, "bold"), anchor="e").pack(anchor="w", pady=(2, 0))

        # Buttons row: Imprimer mobile + Générer A4 local
        btn_row = tk.Frame(panel, bg="white")
        btn_row.pack(fill="x", pady=(8, 2))
        spacer = tk.Frame(btn_row, bg="white")
        spacer.pack(side="left", expand=True)

        def _print_invoice_local(pl=payload, inv_source=inv):
            try:
                # Construire le payload complet via l'enrichissement DB
                pdf_payload = _build_local_payload_from_minimal(inv_source)

                # Créer chemin temporaire dans Documents si possible
                try:
                    doc_dir = os.path.join(os.path.expanduser("~"), "Documents")
                    if not os.path.isdir(doc_dir):
                        doc_dir = tempfile.gettempdir()
                except Exception:
                    doc_dir = tempfile.gettempdir()
                safe_name = f"Facture_{pdf_payload.get('invoice_number','unnamed')}.pdf"
                tmp_path = os.path.join(doc_dir, safe_name)

                # Générer le PDF mobile directement vers tmp_path
                ok, err = generate_invoice_mobile_pdf(pdf_payload, tmp_path)
                if not ok:
                    messagebox.showerror("Erreur impression mobile", err or "Échec génération PDF", parent=popup)
                    return

                # Ouvrir le PDF avec l'application par défaut pour permettre l'impression
                try:
                    _open_file_with_default_app(tmp_path)
                    messagebox.showinfo("Impression mobile", "PDF généré et ouvert pour impression.", parent=popup)
                except Exception:
                    # fallback : tenter impression directe, sinon avertir l'utilisateur du chemin
                    if _print_file_direct(tmp_path):
                        messagebox.showinfo("Impression mobile", "PDF envoyé à l'imprimante.", parent=popup)
                    else:
                        messagebox.showwarning(
                            "Ouverture PDF",
                            f"PDF créé: {tmp_path}\nImpossible d'ouvrir automatiquement, ouvrez-le manuellement pour imprimer.",
                            parent=popup
                        )

            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de générer/imprimer le PDF: {e}", parent=popup)

        b_print = ttk.Button(btn_row, text="Imprimer mobile", command=_print_invoice_local, style="Primary.TButton")
        b_print.pack(side="right", padx=6)

        def _generate_a4(pl=payload):
            try:
                local_id = pl.get("local_facture_id")
                if local_id:
                    _generate_pdf_from_local(int(local_id), popup)
                    return
                # fallback : sauvegarde demandée et génération A4 à partir du payload actuel
                path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                    initialfile=f"Facture_{pl.get('invoice_number','unnamed')}.pdf",
                                                    parent=popup)
                if not path:
                    return
                ok, err = generate_invoice_pdf(pl, path)
                if ok:
                    messagebox.showinfo("Export A4", f"PDF créé: {os.path.basename(path)}", parent=popup)
                else:
                    messagebox.showerror("Erreur export A4", err or "Erreur inconnue", parent=popup)
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de générer le PDF A4: {e}", parent=popup)

        b_a4 = ttk.Button(btn_row, text="Générer PDF local (A4)", command=_generate_a4, style="Default.TButton")
        b_a4.pack(side="right", padx=6)

    # footer info
    info = tk.Label(frame_root, text="Format compact prêt pour impression mobile", font=("Helvetica", 7), bg="white")
    info.pack(side="bottom", pady=(6, 2))

# -----------------------
# Modal EBMS : afficher une invoice renvoyée par getInvoice (structure proche de _view_invoice)
# -----------------------
def _show_invoice_modal_from_ebms(invoice: dict, parent=None):
    """
    Affiche une facture renvoyée par l'EBMS.
    - Totaux calculés à partir des lignes et affichés immédiatement sous le tableau (espacement réduit).
    - Bouton 'Imprimer mobile' génère un PDF via generate_invoice_pdf.
    """
    modal = tk.Toplevel(parent)
    modal.title(f"Facture EBMS: {invoice.get('invoice_number') or invoice.get('invoice_registered_number') or invoice.get('invoice_identifier','')}")
    modal.transient(parent); modal.grab_set()
    width, height = 720, 520  # modal width reduced
    _center_window(modal, parent, width, height)
    frame = tk.Frame(modal, bg="white", padx=10, pady=8); frame.pack(fill="both", expand=True)

    # Header
    tk.Label(frame, text=f"Facture: {invoice.get('invoice_number','')}", font=("Segoe UI",12,"bold"), bg="white").pack(anchor="w")
    tk.Label(frame, text=f"Date: {invoice.get('invoice_date') or invoice.get('invoice_registered_date','')}", bg="white").pack(anchor="w")
    tk.Label(frame, text=f"Identifiant: {invoice.get('invoice_identifier','')}", bg="white").pack(anchor="w")
    tk.Label(frame, text=f"Statut EBMS: {'envoyé' if invoice.get('invoice_registered_number') else invoice.get('facture_statut','')}", bg="white").pack(anchor="w", pady=(0,6))

    # Seller / buyer summary
    seller_lines = []
    for k in ("tp_name","tp_TIN","tp_trade_number","tp_phone_number","tp_address_province","tp_address_commune","tp_address_quartier"):
        v = invoice.get(k)
        if v:
            seller_lines.append(str(v))
    if seller_lines:
        tk.Label(frame, text="Vendeur: " + " / ".join(seller_lines), font=("Segoe UI",9), bg="white", wraplength=680, justify="left").pack(anchor="w", pady=(0,4))

    buyer_lines = []
    for k in ("customer_name","customer_TIN","customer_address","customer_phone_number"):
        v = invoice.get(k)
        if v:
            buyer_lines.append(str(v))
    if buyer_lines:
        tk.Label(frame, text="Client: " + " / ".join(buyer_lines), font=("Segoe UI",9), bg="white", wraplength=680, justify="left").pack(anchor="w", pady=(0,6))

    # Table container (table + summary below) - reduce vertical gaps
    table_container = tk.Frame(frame, bg="white")
    table_container.pack(fill="both", expand=True, pady=(4,2))

    # Items table (top, expands)
    tbl = tk.Frame(table_container, bg="white")
    tbl.pack(side="top", fill="both", expand=True)

    hdrs = ["Désignation","Qté","PU","PV HT","TVA","PU TVAC","Total ligne"]
    for ci, h in enumerate(hdrs):
        tk.Label(tbl, text=h, bg="#eef6ff", font=("Segoe UI",10,"bold"), bd=1, relief="solid", padx=6, pady=6).grid(row=0, column=ci, sticky="nsew")

    items = invoice.get("invoice_items") or invoice.get("items") or []
    # Render rows and accumulate for totals calculation
    row_maps = []
    for ri, it in enumerate(items, start=1):
        des = it.get("item_designation") or it.get("designation") or ""
        qty_v = it.get("item_quantity") or it.get("quantity") or ""
        pu_v = it.get("item_price") or it.get("item_unit_price") or it.get("item_unit_price_used") or it.get("unit_price") or ""
        pu_nvat_v = it.get("item_price_nvat") or it.get("item_unit_price_nvat") or ""
        vat_v = it.get("vat") or it.get("item_vat_amount") or it.get("vat_amount") or ""
        pu_wvat_v = it.get("item_price_wvat") or it.get("item_unit_price_wvat") or ""
        total_v = it.get("item_total_amount") or it.get("line_total_amount") or it.get("item_total") or ""

        vals = [des, str(qty_v), str(pu_v), str(pu_nvat_v), str(vat_v), str(pu_wvat_v), str(total_v)]
        for ci, v in enumerate(vals):
            anchor = "e" if ci >= 1 else "w"
            tk.Label(tbl, text=str(v), bg="white", anchor=anchor, bd=1, relief="solid", padx=6, pady=4).grid(row=ri, column=ci, sticky="nsew")

        row_maps.append({
            "quantity": qty_v,
            "unit_price_used": pu_v,
            "unit_price_nvat": pu_nvat_v,
            "line_total_amount": total_v,
            "vat_amount": vat_v
        })

    # Ensure columns expand reasonably
    for c in range(len(hdrs)):
        tbl.grid_columnconfigure(c, weight=1)

    # Totals calculation (defensive parsing)
    subtotal = 0.0
    vat_sum = 0.0
    for rm in row_maps:
        try:
            qty = float(rm.get("quantity") or 0)
        except Exception:
            try:
                qty = float(str(rm.get("quantity") or 0).replace(",", "."))
            except Exception:
                qty = 0.0
        try:
            pu = float(rm.get("unit_price_used") or rm.get("unit_price_nvat") or rm.get("unit_price") or 0)
        except Exception:
            try:
                pu = float(str(rm.get("unit_price_used") or rm.get("unit_price_nvat") or rm.get("unit_price") or 0).replace(",", "."))
            except Exception:
                pu = 0.0
        # force HT calculation as qty * pu (PVT HTVA)
        try:
            total_line_ht = qty * pu
        except Exception:
            total_line_ht = 0.0
        try:
            vat_amt = float(rm.get("vat_amount") or 0)
        except Exception:
            try:
                vat_amt = float(str(rm.get("vat_amount") or 0).replace(",", "."))
            except Exception:
                vat_amt = 0.0

        subtotal += total_line_ht
        vat_sum += vat_amt

    inv_total_declared = invoice.get("invoice_total") or invoice.get("invoice_total_amount") or invoice.get("grand_total") or invoice.get("total_amount") or invoice.get("invoice_registered_total") or None
    try:
        declared = float(inv_total_declared) if inv_total_declared not in (None, "") else None
    except Exception:
        declared = None
    computed_total = round(subtotal + vat_sum, 2)
    grand_total = declared if declared is not None else computed_total

    # Display totals directly under the table with minimal spacing
    totals_display_frame = tk.Frame(table_container, bg="white")
    totals_display_frame.pack(fill="x", pady=(2,0))  # smaller gap to table

    totals_left = tk.Frame(totals_display_frame, bg="white")
    totals_left.pack(side="left", anchor="w", padx=0, pady=0)

    # reduce internal vertical spacing on labels (pady=0)
    tk.Label(totals_left, text=f"PVT HTVA: {subtotal:.2f}", bg="white", anchor="w", justify="left", font=("Segoe UI",10)).pack(anchor="w", pady=0)
    tk.Label(totals_left, text=f"TVA: {vat_sum:.2f}", bg="white", anchor="w", justify="left", font=("Segoe UI",10)).pack(anchor="w", pady=0)
    tk.Label(totals_left, text=f"Total TVAC: {grand_total:.2f}", font=("Segoe UI",10,"bold"), bg="white", anchor="w", justify="left").pack(anchor="w", pady=(0,0))

    # Right side: computed electronic signature if available (compact)
    invoice_identifier_val = invoice.get("invoice_identifier") or invoice.get("invoice_registered_number") or invoice.get("invoice_number") or ""
    electronic_signature_calc = _sha1_hex_normalized(invoice_identifier_val)
    if electronic_signature_calc:
        sig_frame = tk.Frame(totals_display_frame, bg="white")
        sig_frame.pack(side="right", anchor="e", padx=0, pady=0)
        tk.Label(sig_frame, text=f"Signature calculée: {electronic_signature_calc}", bg="white", wraplength=300, justify="right", font=("Segoe UI",8)).pack(anchor="e", pady=0)

        # Helper to build PDF payload
        def _build_pdf_payload_from_ebms(inv: dict) -> dict:
            tp = {
                "tp_name": inv.get("tp_name") or inv.get("seller_name") or inv.get("tp_trade_name") or "",
                "tp_TIN": inv.get("tp_TIN") or inv.get("tp_tin") or "",
                "tp_fiscal_center": inv.get("tp_fiscal_center") or "",
                "tp_trade_number": inv.get("tp_trade_number") or "",
                "tp_postal_number": inv.get("tp_postal_number") or "",
                "tp_phone_number": inv.get("tp_phone_number") or inv.get("tp_phone") or "",
                "tp_address_commune": inv.get("tp_address_commune") or "",
                "tp_address_quartier": inv.get("tp_address_quartier") or "",
                "tp_address_avenue": inv.get("tp_address_avenue") or "",
                "tp_address_rue": inv.get("tp_address_rue") or "",
                "tp_legal_form": inv.get("tp_legal_form") or "",
                "tp_activity_sector": inv.get("tp_activity_sector") or "",
                "vat_taxpayer": inv.get("vat_taxpayer") or inv.get("tp_vat_taxpayer") or 0
            }
            client = {
                "customer_name": inv.get("buyer_name") or inv.get("customer_name") or inv.get("client_name") or "",
                "customer_TIN": inv.get("buyer_tin") or inv.get("customer_TIN") or "",
                "customer_address": inv.get("buyer_address") or inv.get("customer_address") or "",
                "customer_phone_number": inv.get("buyer_phone") or inv.get("customer_phone_number") or "",
                "vat_customer_payer": inv.get("vat_customer_payer") or inv.get("vat_customer") or 0

            }
            items_raw = inv.get("invoice_items") or inv.get("items") or []
            items_out = []
            for it in items_raw:
                items_out.append({
                    "item_designation": it.get("item_designation") or it.get("designation") or it.get("name") or "",
                    "item_quantity": it.get("item_quantity") or it.get("quantity") or it.get("qty") or 0,
                    "item_unit_price": it.get("item_price") or it.get("item_unit_price") or it.get("unit_price") or 0,
                    "item_price_nvat": it.get("item_price_nvat") or it.get("item_unit_price_nvat") or "",
                    "item_price_wvat": it.get("item_price_wvat") or it.get("item_unit_price_wvat") or "",
                    "item_total_amount": it.get("item_total_amount") or it.get("line_total_amount") or it.get("total") or 0,
                    "item_vat_amount": it.get("vat") or it.get("item_vat_amount") or it.get("vat_amount") or 0
                })
            # compute totals defensively
            subtotal = 0.0
            vat_sum = 0.0
            for it in items_out:
                try:
                    q = float(it.get("item_quantity") or 0)
                except Exception:
                    q = 0.0
                try:
                    pu = float(it.get("item_unit_price") or 0)
                except Exception:
                    pu = 0.0
                subtotal += q * pu
                try:
                    vat_sum += float(it.get("item_vat_amount") or 0)
                except Exception:
                    vat_sum += 0.0
            totals = {"subtotal": round(subtotal, 2), "vat": round(vat_sum, 2), "grand_total": round(subtotal + vat_sum, 2)}

            pdf_payload = {
                "invoice_number": inv.get("invoice_registered_number") or inv.get("invoice_number") or inv.get("invoice_ref") or "",
                "invoice_date": inv.get("invoice_registered_date") or inv.get("invoice_date") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "invoice_identifier": inv.get("invoice_identifier") or "",
                "tp": tp,
                "client": client,
                "invoice_items": items_out,
                "totals": totals,
                "electronic_signature": inv.get("invoice_identifier") or inv.get("electronic_signature") or inv.get("signature") or ""
            }
            return pdf_payload

        # Print mobile handler
        def _print_mobile(inv: dict):
            try:
                pdf_payload = _build_pdf_payload_from_ebms(inv)
                path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=f"Facture_{pdf_payload['invoice_number'] or 'unnamed'}.pdf", parent=modal)
                if not path:
                    return
                ok, err = generate_invoice_mobile_pdf(pdf_payload, path)
                if ok:
                    messagebox.showinfo("Impression mobile", f"PDF créé: {os.path.basename(path)}", parent=modal)
                else:
                    messagebox.showerror("Erreur PDF", str(err), parent=modal)
            except Exception as e:
                logger.exception("Erreur impression mobile: %s", e)
                messagebox.showerror("Erreur", f"Impossible de générer le PDF: {e}", parent=modal)

        # Buttons row (print mobile + close)
        btn_row = tk.Frame(frame, bg="white"); btn_row.pack(fill="x", pady=(8,0))
        spacer = tk.Frame(btn_row, bg="white"); spacer.pack(side="left", expand=True)

        b_print_mobile = ttk.Button(btn_row, text="Imprimer mobile", command=lambda inv=invoice: _print_mobile(inv), style="Primary.TButton")
        b_print_mobile.pack(side="right", padx=6)
        def _close(): modal.destroy()
        b_close = ttk.Button(btn_row, text="Fermer", command=_close, style="Default.TButton"); b_close.pack(side="right", padx=6)


# -----------------------
# getInvoice : appel EBMS puis fallback local (remplacée)
# -----------------------
def get_invoice_details(invoice_identifier: str, parent=None):
    """
    Appelle EBMS getInvoice ; si succès, ouvre une modal détaillée avec la première invoice renvoyée.
    En cas d'erreur réseau / réponse négative, utilise le fallback local et ouvre la modal
    imprimable (_show_invoices_modal_printable) en mode source_is_ebms=False.
    """
    token = None
    try:
        token = obtenir_token_auto()
    except Exception:
        token = None

    # Try EBMS first if token available
    if token:
        try:
            url = "https://ebms.obr.gov.bi:9443/ebms_api/getInvoice/"
            headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
            payload = {"invoice_identifier": invoice_identifier}
            resp = requests.post(url, json=payload, headers=headers, verify=False, timeout=30)
            j = None
            try:
                j = resp.json()
            except Exception:
                j = None

            if resp is not None and resp.status_code == 200 and isinstance(j, dict) and j.get("success"):
                invoices = []
                if isinstance(j.get("result"), dict):
                    if isinstance(j["result"].get("invoices"), list):
                        invoices = j["result"]["invoices"]
                    elif isinstance(j["result"].get("invoice"), dict):
                        invoices = [j["result"]["invoice"]]
                    else:
                        invoices = [j["result"]]
                elif isinstance(j.get("invoices"), list):
                    invoices = j["invoices"]
                elif isinstance(j.get("invoices"), dict):
                    invoices = [j["invoices"]]
                elif isinstance(j.get("result"), list):
                    invoices = j["result"]

                if isinstance(invoices, dict):
                    invoices = [invoices]
                if invoices:
                    try:
                        _show_invoice_modal_from_ebms(invoices[0], parent=parent)
                    except Exception:
                        logger.exception("Affichage modal EBMS a échoué")
                        messagebox.showerror("Erreur", "Impossible d'afficher la facture renvoyée par l'EBMS.", parent=parent)
                    return
                else:
                    messagebox.showinfo("GetInvoice", "Aucune facture renvoyée par l'EBMS pour cet identifiant.", parent=parent)
            else:
                try:
                    body = resp.text if resp is not None else str(j)
                except Exception:
                    body = str(j)
                messagebox.showerror("Erreur GetInvoice", body, parent=parent)
        except Exception:
            logger.exception("Erreur réseau GetInvoice")

    # Fallback local reconstruction : récupère la facture locale et ouvre la modal imprimable
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT f.id, f.invoice_number, f.invoice_date, f.invoice_identifier, f.facture_statut,
                   COALESCE(c.customer_name, '') as customer_name, COALESCE(c.customer_TIN, '') as customer_TIN,
                   COALESCE(c.customer_address, '') as customer_address, COALESCE(c.customer_phone_number,'') as customer_phone
            FROM facture f LEFT JOIN client c ON f.client_id = c.id
            WHERE f.invoice_identifier = ? OR f.invoice_number = ?
            LIMIT 1
        """, (invoice_identifier, invoice_identifier))
        fr = cur.fetchone()
        if not fr:
            conn.close()
            # show minimal printable modal so user can still generate an invoice
            _show_invoices_modal_printable([{
                "invoice_number": invoice_identifier,
                "invoice_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "items": []
            }], parent=parent, source_is_ebms=False)
            return

        # Unpack safely
        fid = fr[0]; inv_num = fr[1]; inv_date = fr[2]; inv_ident = fr[3]
        cust_name = fr[5]; cust_tin = fr[6]; cust_addr = fr[7]; cust_phone = fr[8]

        # Build minimal payload for the printable modal: _show_invoices_modal_printable will enrich if possible
        payload = {
            "invoice_number": inv_num,
            "invoice_date": inv_date,
            "invoice_identifier": inv_ident,
            "buyer_name": cust_name,
            "buyer_tin": cust_tin,
            "buyer_address": cust_addr,
            "buyer_phone": cust_phone,
            "items": []  # will be enriched by _show_invoices_modal_printable via DB if available
        }
        conn.close()

        _show_invoices_modal_printable([payload], parent=parent, source_is_ebms=False)
    except Exception:
        logger.exception("Fallback local getInvoice failed")
        messagebox.showerror("Erreur", "Impossible de récupérer la facture.", parent=parent)


# -----------------------
# Vue locale simple de la facture
# -----------------------
def _view_invoice(row_data: Dict, parent=None):
    modal = tk.Toplevel(parent); modal.title(f"Facture: {row_data.get('invoice_number','')}"); modal.transient(parent); modal.grab_set()
    width, height = 620, 380; _center_window(modal, parent, width, height)
    frame = tk.Frame(modal, bg="white", padx=8, pady=8); frame.pack(fill="both", expand=True)
    tk.Label(frame, text=f"Facture: {row_data.get('invoice_number')}", font=("Segoe UI",12,"bold"), bg="white").pack(anchor="w")
    tk.Label(frame, text=f"Date: {row_data.get('invoice_date')}", bg="white").pack(anchor="w")
    tk.Label(frame, text=f"Client: {row_data.get('customer_name')}", bg="white").pack(anchor="w")
    tk.Label(frame, text=f"Statut: {row_data.get('facture_statut')}", bg="white").pack(anchor="w", pady=(0,8))
    try:
        conn = get_connection(); cur = conn.cursor()
        # items pour l’aperçu simple
        cur.execute("SELECT item_designation, quantity, unit_price_used, unit_price_nvat, vat_amount, line_total_amount FROM article WHERE facture_id=?", (row_data.get("id"),))
        arts = cur.fetchall() or []; conn.close()
    except Exception:
        arts = []
    tbl = tk.Frame(frame, bg="white"); tbl.pack(fill="both", expand=True)
    hdrs = ["Désignation","Qté","PU utilisé","PU HT","TVA","Total ligne"]
    for ci, h in enumerate(hdrs):
        tk.Label(tbl, text=h, bg="#eef6ff", font=("Segoe UI",10,"bold"), bd=1, relief="solid", padx=6, pady=6).grid(row=0, column=ci, sticky="nsew")
    for ri, a in enumerate(arts, start=1):
        des, qty, pu_used, pu_nvat, vat_amt, total = a
        vals = [
            des or "",
            f"{float(qty or 0):.2f}",
            f"{float(pu_used or 0):.2f}",
            f"{float(pu_nvat or 0):.2f}",
            f"{float(vat_amt or 0):.2f}",
            f"{float(total or 0):.2f}",
        ]
        for ci, v in enumerate(vals):
            anchor = "e" if ci >= 1 else "w"
            tk.Label(tbl, text=str(v), bg=CARD_BG, anchor=anchor, bd=1, relief="solid", padx=6, pady=6).grid(row=ri, column=ci, sticky="nsew")
    _modal_buttons(frame, [("Générer PDF local (A4)", lambda: _generate_pdf_from_local(row_data.get("id"), modal), "Primary"), ("Fermer", modal.destroy, "Default")], align="right")


# -----------------------
# Génération PDF local (wrapper)
# -----------------------
def _generate_pdf_from_local(facture_id: int, parent_win):
    """
    Lit facture + contribuable + client + articles depuis la DB, construit payload
    et appelle generate_invoice_pdf. Affiche des messagebox pour retour utilisateur.
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, invoice_number, invoice_date, invoice_identifier, contribuable_id, client_id, total_amount "
            "FROM facture WHERE id = ?",
            (facture_id,)
        )
        fr = cur.fetchone()
        if not fr:
            messagebox.showerror("Erreur", "Facture introuvable", parent=parent_win)
            conn.close()
            return

        fr_t = tuple(fr)
        fid = fr_t[0] if len(fr_t) > 0 else None
        inv_num = fr_t[1] if len(fr_t) > 1 else ""
        inv_date = fr_t[2] if len(fr_t) > 2 else ""
        inv_ident = fr_t[3] if len(fr_t) > 3 else ""
        tp_id = fr_t[4] if len(fr_t) > 4 else None
        cl_id = fr_t[5] if len(fr_t) > 5 else None
        inv_total_declared = fr_t[6] if len(fr_t) > 6 else None

        # contribuable
        tp = {}
        if tp_id:
            try:
                cur.execute("SELECT * FROM contribuable WHERE id = ? LIMIT 1", (tp_id,))
                rtp = cur.fetchone()
                if rtp:
                    cols = [d[0] for d in cur.description]
                    tp = dict(zip(cols, tuple(rtp)))
            except Exception:
                tp = {}

        # client
        client = {}
        if cl_id:
            try:
                cur.execute("SELECT * FROM client WHERE id = ? LIMIT 1", (cl_id,))
                rc = cur.fetchone()
                if rc:
                    cols = [d[0] for d in cur.description]
                    client = dict(zip(cols, tuple(rc)))
            except Exception:
                client = {}

        # articles : sélection colonnes présentes
        cur.execute("PRAGMA table_info(article)")
        cols_info = cur.fetchall() or []
        existing_cols = {c[1] for c in cols_info}
        wanted = ["item_designation", "quantity", "unit_price_used", "unit_price_nvat", "vat_amount", "line_total_amount"]
        select_cols = [c for c in wanted if c in existing_cols]

        items = []
        subtotal = 0.0
        vat_sum = 0.0
        if select_cols:
            sel = ", ".join(select_cols)
            cur.execute(f"SELECT {sel} FROM article WHERE facture_id = ? ORDER BY id ASC", (facture_id,))
            arts = cur.fetchall() or []
            for a in arts:
                row_map = {select_cols[i]: (a[i] if i < len(a) else None) for i in range(len(select_cols))}
                # normalize
                try:
                    qty = float(row_map.get("quantity") or 0)
                except Exception:
                    try: qty = float(str(row_map.get("quantity")).replace(",", "."))
                    except: qty = 0.0
                try:
                    pu = float(row_map.get("unit_price_used") or row_map.get("unit_price_nvat") or 0)
                except Exception:
                    try: pu = float(str(row_map.get("unit_price_used") or row_map.get("unit_price_nvat")).replace(",", "."))
                    except: pu = 0.0
                try:
                    total_line = float(row_map.get("line_total_amount") or qty * pu)
                except Exception:
                    total_line = qty * pu
                try:
                    vat_amt = float(row_map.get("vat_amount") or 0)
                except Exception:
                    try: vat_amt = float(str(row_map.get("vat_amount")).replace(",", "."))
                    except: vat_amt = 0.0

                subtotal += qty * pu
                vat_sum += vat_amt

                items.append({
                    "item_designation": row_map.get("item_designation") or "",
                    "item_quantity": qty,
                    "item_unit_price": pu,
                    "item_total_amount": total_line,
                    "item_vat_amount": vat_amt
                })

        conn.close()

    except Exception as e:
        logger.exception("PDF from local failed: %s", e)
        messagebox.showerror("Erreur", str(e), parent=parent_win)
        return

    # totals
    try:
        declared = float(inv_total_declared) if inv_total_declared not in (None, "") else None
    except Exception:
        declared = None
    computed_total = round(subtotal + vat_sum, 2)
    grand_total = declared if declared is not None else computed_total

    payload = {
        "invoice_number": inv_num,
        "invoice_date": inv_date,
        "invoice_identifier": inv_ident,
        "tp": {
            "tp_name": tp.get("tp_name") or tp.get("name") or "",
            "tp_TIN": tp.get("tp_TIN") or tp.get("tin") or "",
            "tp_fiscal_center": tp.get("tp_fiscal_center") or "",
            "tp_trade_number": tp.get("tp_trade_number") or "",
            "tp_postal_number": tp.get("tp_postal_number") or "",
            "tp_phone_number": tp.get("tp_phone_number") or "",
            "tp_address_commune": tp.get("tp_address_commune") or "",
            "tp_address_quartier": tp.get("tp_address_quartier") or "",
            "tp_address_avenue": tp.get("tp_address_avenue") or "",
            "tp_address_rue": tp.get("tp_address_rue") or "",
            "tp_legal_form": tp.get("tp_legal_form") or "",
            "tp_activity_sector": tp.get("tp_activity_sector") or "",
            "vat_taxpayer": tp.get("vat_taxpayer") or 0
        },
        "client": {
            "customer_name": client.get("customer_name") or client.get("name") or "",
            "customer_TIN": client.get("customer_TIN") or client.get("tin") or "",
            "customer_address": client.get("customer_address") or client.get("address") or "",
            "vat_customer_payer": client.get("vat_customer_payer") or 0
        },
        "invoice_items": items,
        "totals": {
            "subtotal": round(subtotal, 2),
            "vat": round(vat_sum, 2),
            "grand_total": round(grand_total, 2)
        },
        "footer_notes": [
            "* Mention obligatoire",
            "Les personnes assujetties à la TVA ne remplissent pas les deux dernières lignes"
        ],
        "electronic_signature": inv_ident or ""
    }

    # save dialog and generate
    path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=f"Facture_{inv_num or 'unnamed'}.pdf", parent=parent_win)
    if not path:
        return

    ok, err = generate_invoice_pdf(payload, path)
    if ok:
        messagebox.showinfo("Export PDF", f"PDF créé: {os.path.basename(path)}", parent=parent_win)
    else:
        messagebox.showerror("Erreur export PDF", err, parent=parent_win)


def _build_local_payload_from_minimal(inv: dict) -> dict:
    """
    Enrichit une facture minimale (inv) en lisant la base locale si possible.
    Retourne un payload structuré prêt pour generate_invoice_pdf / generate_invoice_mobile_pdf.
    Copie la logique défensive de _generate_pdf_from_local afin d'être cohérent.
    """
    payload = {
        "invoice_number": inv.get("invoice_number") or inv.get("invoice_registered_number") or inv.get("invoice_identifier") or "",
        "invoice_date": inv.get("invoice_date") or inv.get("invoice_registered_date") or "",
        "invoice_identifier": inv.get("invoice_identifier") or "",
        "tp": {},
        "client": {},
        "invoice_items": inv.get("invoice_items") or inv.get("items") or [],
        "totals": inv.get("totals") or {},
        "electronic_signature": inv.get("electronic_signature") or inv.get("signature") or "",
        "local_facture_id": None
    }

    inv_ident = payload["invoice_identifier"] or payload["invoice_number"]

    # If inv already contains full details, merge and return
    if (payload.get("invoice_items") and (inv.get("tp") or inv.get("client"))):
        if inv.get("tp"):
            payload["tp"].update(inv.get("tp") or {})
        if inv.get("client"):
            payload["client"].update(inv.get("client") or {})
        return payload

    if not inv_ident:
        return payload

    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT f.id, f.invoice_number, f.invoice_date, f.invoice_identifier, f.contribuable_id, f.client_id, f.total_amount
            FROM facture f
            WHERE f.invoice_identifier = ? OR f.invoice_number = ?
            LIMIT 1
        """, (inv_ident, inv_ident))
        fr = cur.fetchone()
        if not fr:
            conn.close()
            return payload

        fid = fr[0]
        inv_num = fr[1] or ""
        inv_date = fr[2] or ""
        inv_ident_db = fr[3] or ""
        tp_id = fr[4]
        cl_id = fr[5]
        inv_total_declared = fr[6] if len(fr) > 6 else None

        # contribuable (tp)
        tp = {}
        if tp_id:
            try:
                cur.execute("SELECT * FROM contribuable WHERE id = ? LIMIT 1", (tp_id,))
                rtp = cur.fetchone()
                if rtp:
                    cols = [d[0] for d in cur.description]
                    tp = dict(zip(cols, tuple(rtp)))
            except Exception:
                tp = {}

        # client
        client = {}
        if cl_id:
            try:
                cur.execute("SELECT * FROM client WHERE id = ? LIMIT 1", (cl_id,))
                rc = cur.fetchone()
                if rc:
                    cols = [d[0] for d in cur.description]
                    client = dict(zip(cols, tuple(rc)))
            except Exception:
                client = {}

        # articles : choix des colonnes présentes
        cur.execute("PRAGMA table_info(article)")
        cols_info = cur.fetchall() or []
        existing_cols = {c[1] for c in cols_info}
        wanted = ["item_designation", "quantity", "unit_price_used", "unit_price_nvat", "vat_amount", "line_total_amount"]
        select_cols = [c for c in wanted if c in existing_cols]

        items = []
        subtotal = 0.0
        vat_sum = 0.0

        def _to_float_safe(x):
            try:
                if x is None or x == "":
                    return 0.0
                if isinstance(x, (int, float)):
                    return float(x)
                s = str(x).strip().replace(" ", "").replace("\u00A0", "")
                if s.count(",") > 0 and s.count(".") == 0:
                    s = s.replace(",", ".")
                else:
                    s = s.replace(",", "")
                return float(s)
            except Exception:
                return 0.0

        if select_cols:
            sel = ", ".join(select_cols)
            cur.execute(f"SELECT {sel} FROM article WHERE facture_id = ? ORDER BY id ASC", (fid,))
            arts = cur.fetchall() or []
            for a in arts:
                row_map = {select_cols[i]: (a[i] if i < len(a) else None) for i in range(len(select_cols))}
                qty = _to_float_safe(row_map.get("quantity"))
                pu_used = _to_float_safe(row_map.get("unit_price_used"))
                pu_nvat = _to_float_safe(row_map.get("unit_price_nvat"))
                line_total = _to_float_safe(row_map.get("line_total_amount")) or round(qty * (pu_used or pu_nvat or 0.0), 6)
                vat_amt = _to_float_safe(row_map.get("vat_amount"))

                subtotal += qty * (pu_used or pu_nvat or 0.0)
                vat_sum += vat_amt

                items.append({
                    "item_designation": row_map.get("item_designation") or "",
                    "item_quantity": qty,
                    "item_unit_price": pu_used or pu_nvat,
                    "item_total_amount": line_total,
                    "item_vat_amount": vat_amt
                })

        conn.close()

        # totals: prefer declared if present
        try:
            declared = float(inv_total_declared) if inv_total_declared not in (None, "") else None
        except Exception:
            declared = None
        computed_total = round(subtotal + vat_sum, 2)
        grand_total = declared if declared is not None else computed_total

        # build payload
        payload["invoice_number"] = inv_num or payload["invoice_number"]
        payload["invoice_date"] = inv_date or payload["invoice_date"]
        payload["invoice_identifier"] = inv_ident_db or payload["invoice_identifier"]
        payload["tp"] = {
            "tp_name": tp.get("tp_name") or tp.get("name") or "",
            "tp_TIN": tp.get("tp_TIN") or tp.get("tin") or "",
            "tp_fiscal_center": tp.get("tp_fiscal_center") or "",
            "tp_trade_number": tp.get("tp_trade_number") or "",
            "tp_postal_number": tp.get("tp_postal_number") or "",
            "tp_phone_number": tp.get("tp_phone_number") or "",
            "tp_address_commune": tp.get("tp_address_commune") or "",
            "tp_address_quartier": tp.get("tp_address_quartier") or "",
            "tp_address_avenue": tp.get("tp_address_avenue") or "",
            "tp_address_rue": tp.get("tp_address_rue") or "",
            "tp_address_number": tp.get("tp_address_number") or "",
            "tp_legal_form": tp.get("tp_legal_form") or "",
            "tp_activity_sector": tp.get("tp_activity_sector") or "",
            "vat_taxpayer": tp.get("vat_taxpayer") or 0
        }
        payload["client"] = {
            "customer_name": client.get("customer_name") or client.get("name") or "",
            "customer_TIN": client.get("customer_TIN") or client.get("tin") or "",
            "customer_address": client.get("customer_address") or client.get("address") or "",
            "customer_phone_number": client.get("customer_phone_number") or client.get("phone_number") or "",
            "vat_customer_payer": client.get("vat_customer_payer") or 0
        }
        payload["invoice_items"] = items
        payload["totals"] = {
            "subtotal": round(subtotal, 2),
            "vat": round(vat_sum, 2),
            "grand_total": round(grand_total, 2)
        }
        payload["electronic_signature"] = payload.get("electronic_signature") or tp.get("electronic_signature") or ""
        payload["local_facture_id"] = fid

        return payload

    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        return payload


# -----------------------
# retry_invoice_local_and_send (Réessayer)
# -----------------------
def retry_invoice_local_and_send(facture_id: int, parent=None):
    """
    Reconstitue la facture locale et renvoie vers l'API OBR (addInvoice).
    - FORCER invoice_identifier maintenant (ne pas utiliser la valeur en table) en forme:
      {tin}/{system_id}/{YYYYMMDDhhmmss}/{invoice_number}
    - N'écrit invoice_identifier local que si l'appel addInvoice réussit.
    - Génère invoice_signature et electronic_signature (SHA-256) au moment de l'envoi.
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, invoice_number, invoice_date, invoice_type, invoice_identifier,
                   payment_type, currency, cancelled_invoice_ref, cn_motif,
                   contribuable_id, client_id, total_amount
            FROM facture WHERE id=?
        """, (facture_id,))
        row = cur.fetchone()
        if not row:
            messagebox.showerror("Erreur", "Facture introuvable.", parent=parent)
            conn.close()
            return

        # safe unpack and pad
        row_t = tuple(row)
        (fid, inv_num, inv_date, inv_type, inv_ident, pay_type, curr,
         cancel_ref, cn_motif, tp_id, cl_id, inv_total_declared) = (row_t + (None,) * 12)[:12]

        # load contribuable snapshot
        tp_snapshot = {}
        if tp_id:
            try:
                cur.execute("SELECT * FROM contribuable WHERE id=? LIMIT 1", (tp_id,))
                rtp = cur.fetchone()
                if rtp:
                    cols = [d[0] for d in cur.description]
                    tp_snapshot = dict(zip(cols, tuple(rtp)))
            except Exception:
                logger.exception("Load contribuable failed")

        # load client snapshot
        cl_snapshot = {}
        if cl_id:
            try:
                cur.execute("SELECT * FROM client WHERE id=? LIMIT 1", (cl_id,))
                rc = cur.fetchone()
                if rc:
                    cols = [d[0] for d in cur.description]
                    cl_snapshot = dict(zip(cols, tuple(rc)))
            except Exception:
                logger.exception("Load client failed")

        # determine article columns present
        cur.execute("PRAGMA table_info(article)")
        cols_info = cur.fetchall() or []
        existing_cols = {c[1] for c in cols_info}
        wanted = [
            "id", "item_code", "item_designation", "quantity",
            "unit_price_used", "unit_price_nvat", "vat_amount", "unit_price_wvat", "line_total_amount"
        ]
        select_cols = [c for c in wanted if c in existing_cols]
        if not select_cols:
            fallback = ["id", "item_code", "item_designation", "quantity", "line_total_amount"]
            select_cols = [c for c in fallback if c in existing_cols]

        select_expr = ", ".join(select_cols) if select_cols else "id"
        cur.execute(f"SELECT {select_expr} FROM article WHERE facture_id=? ORDER BY id ASC", (facture_id,))
        arts = cur.fetchall() or []

        # helper to parse floats defensively
        def _parse_float(x):
            try:
                if x is None or x == "":
                    return 0.0
                if isinstance(x, (int, float)):
                    return float(x)
                s = str(x).strip().replace(" ", "").replace("\u00A0", "")
                # comma decimal handling
                if s.count(",") > 0 and s.count(".") == 0:
                    s = s.replace(",", ".")
                else:
                    s = s.replace(",", "")
                return float(s)
            except Exception:
                return 0.0

        lignes = []
        subtotal = 0.0
        vat_sum = 0.0
        for a in arts:
            row_map = {select_cols[i]: (a[i] if i < len(a) else None) for i in range(len(select_cols))}

            qty = _parse_float(row_map.get("quantity", 0))
            pu_used = _parse_float(row_map.get("unit_price_used", 0))
            pu_nvat = _parse_float(row_map.get("unit_price_nvat", 0))
            pu_wvat = _parse_float(row_map.get("unit_price_wvat", 0)) if "unit_price_wvat" in row_map else 0.0

            # choose best unit price (prefer used then nvat then wvat)
            unit_price = pu_used if pu_used else (pu_nvat if pu_nvat else pu_wvat)

            # line total
            total_raw = row_map.get("line_total_amount", None)
            if total_raw in (None, ""):
                total = round(unit_price * qty, 6)
            else:
                total = round(_parse_float(total_raw), 6)

            # vat amount (line)
            vat_amt = _parse_float(row_map.get("vat_amount", 0))
            # infer VAT if missing but have HT and WVAT
            if vat_amt == 0.0 and pu_nvat and pu_wvat:
                vat_per_unit = pu_wvat - pu_nvat
                vat_amt = round(vat_per_unit * qty, 6)
            subtotal += total
            vat_sum += vat_amt

            # Ensure presence of all required keys expected by OBR
            item_price_nvat = pu_nvat if pu_nvat else unit_price
            item_price_wvat = pu_wvat if pu_wvat else round(item_price_nvat + (vat_amt / qty) if qty else item_price_nvat, 6)
            item_price = unit_price

            ligne = {
                "item_designation": row_map.get("item_designation") or "",
                "item_code": row_map.get("item_code") or "",
                "item_quantity": str(qty),
                "item_unit_price": str(round(unit_price, 6)),
                "item_unit_price_nvat": str(round(item_price_nvat, 6)),
                "item_unit_price_wvat": str(round(item_price_wvat, 6)),
                "item_price": str(round(item_price, 6)),
                "item_price_nvat": str(round(item_price_nvat, 6)),
                "vat": str(round(vat_amt, 6)),
                "item_price_wvat": str(round(item_price_wvat, 6)),
                "item_total_amount": str(round(total, 6)),
                "item_vat_amount": str(round(vat_amt, 6)),
            }
            lignes.append(ligne)

        # close read cursor/connection before send
        conn.close()
    except Exception as e:
        logger.exception("Reconstruct local invoice failed: %s", e)
        messagebox.showerror("Erreur", f"Impossible de reconstruire la facture: {e}", parent=parent)
        return

    # FORCER invoice_identifier now (do not use existing): {tin}/{system_id}/{YYYYMMDDhhmmss}/{invoice_number}
    try:
        tin = (tp_snapshot.get("tp_TIN") or tp_snapshot.get("tp_tin") or "").strip() or "UNKNOWN"
        system_id = str(get_system_id() or "").strip() or "UNKNOWN"
        ts_compact = datetime.now().strftime("%Y%m%d%H%M%S")
        forced_invoice_identifier = f"{tin}/{system_id}/{ts_compact}/{inv_num or ''}"
    except Exception:
        forced_invoice_identifier = inv_ident or ""

    # FORCER la signature maintenant (do not read from table)
    sig_dt_ui = None
    try:
        if inv_date:
            dt_try = datetime.strptime(str(inv_date).strip(), "%Y-%m-%d %H:%M:%S")
            if dt_try <= datetime.now():
                sig_dt_ui = dt_try.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        sig_dt_ui = None

    invoice_signature, sig_date_field, electronic_signature = _build_obr_invoice_signature(tp_snapshot or {}, inv_num, sig_dt_ui=sig_dt_ui)

    invoice_payload = {
        "invoice_number": inv_num,
        "invoice_date": inv_date or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "invoice_identifier": forced_invoice_identifier,
        "invoice_type": (inv_type or "").split(" -")[0] if isinstance(inv_type, str) else (inv_type or ""),
        "tp_type": tp_snapshot.get("tp_type","") or tp_snapshot.get("type",""),
        "tp_name": tp_snapshot.get("tp_name","") or tp_snapshot.get("tp_trade_name",""),
        "tp_TIN": tp_snapshot.get("tp_TIN","") or tp_snapshot.get("tp_tin",""),
        "tp_trade_number": tp_snapshot.get("tp_trade_number",""),
        "tp_phone_number": tp_snapshot.get("tp_phone_number","") or tp_snapshot.get("phone",""),
        "tp_address_province": tp_snapshot.get("tp_address_province",""),
        "tp_address_commune": tp_snapshot.get("tp_address_commune",""),
        "tp_address_quartier": tp_snapshot.get("tp_address_quartier",""),
        "tp_address_avenue": tp_snapshot.get("tp_address_avenue",""),
        "tp_address_number": tp_snapshot.get("tp_address_number",""),
        "tp_fiscal_center": tp_snapshot.get("tp_fiscal_center",""),
        "tp_legal_form": tp_snapshot.get("tp_legal_form",""),
        "tp_activity_sector": tp_snapshot.get("tp_activity_sector",""),
        "vat_taxpayer": tp_snapshot.get("vat_taxpayer",""),
        "ct_taxpayer": tp_snapshot.get("ct_taxpayer",""),
        "tl_taxpayer": tp_snapshot.get("tl_taxpayer",""),
        "customer_name": cl_snapshot.get("customer_name",""),
        "customer_TIN": cl_snapshot.get("customer_TIN",""),
        "customer_address": cl_snapshot.get("customer_address",""),
        "customer_phone_number": cl_snapshot.get("customer_phone_number",""),
        "vat_customer_payer": "1" if (cl_snapshot.get("vat_customer_payer") or "") else "",
        "payment_type": pay_type or "",
        "invoice_ref": inv_num or "",
        "invoice_currency": curr or "BIF",
        "invoice_items": lignes,
        "invoice_signature": invoice_signature,
        "invoice_signature_date": sig_date_field,
        "electronic_signature": electronic_signature,
        "system_or_device_id": system_id
    }

    token = obtenir_token_auto()
    if not token:
        messagebox.showerror("Erreur", "Token OBR introuvable", parent=parent)
        return

    try:
        resp = requests.post(
            "https://ebms.obr.gov.bi:9443/ebms_api/addInvoice/",
            json=invoice_payload,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            verify=False,
            timeout=30
        )
    except Exception as ex:
        logger.exception("Network addInvoice failed: %s", ex)
        messagebox.showerror("Erreur réseau", f"Échec envoi: {ex}", parent=parent)
        return

    try:
        j = resp.json() if resp is not None else {}
    except Exception:
        j = {}

    if resp is not None and resp.status_code == 200 and j.get("success"):
        result = j.get("result") or {}
        try:
            conn2 = get_connection()
            cur2 = conn2.cursor()
            # update local status and save signature info and invoice_identifier
            cur2.execute(
                "UPDATE facture SET facture_statut='envoyé', invoice_signature=?, invoice_signature_date=?, electronic_signature=?, invoice_identifier=? WHERE id=?",
                (invoice_signature, sig_date_field, electronic_signature, forced_invoice_identifier, facture_id)
            )
            try:
                cur2.execute("""
                    INSERT INTO accuse_reception (invoice_registered_number, invoice_registered_date, electronic_signature, facture_id)
                    VALUES (?,?,?,?)
                """, (
                    result.get("invoice_registered_number") or invoice_payload.get("invoice_number"),
                    result.get("invoice_registered_date") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    result.get("electronic_signature") or invoice_payload.get("electronic_signature"),
                    facture_id
                ))
            except Exception:
                logger.exception("Insert accuse_reception failed")
            conn2.commit()
            conn2.close()
        except Exception:
            logger.exception("Update local after addInvoice failed")

        messagebox.showinfo("Succès", "Facture renvoyée avec succès.", parent=parent)
        return
    else:
        try:
            body = resp.text if resp is not None else str(j)
        except Exception:
            body = str(j)
        messagebox.showerror("Erreur OBR", body, parent=parent)
        return


# -----------------------
# Modal annulation -> motif + cancelInvoice
# -----------------------
def _open_cancel_modal_with_send(invoice_identifier: str, facture_id_local: int, parent_win=None):
    if not invoice_identifier:
        messagebox.showerror("Erreur", "Identifiant OBR introuvable.", parent=parent_win)
        return
    modal = tk.Toplevel(parent_win)
    modal.title("Annuler facture")
    modal.transient(parent_win); modal.grab_set()
    width, height = 490, 260
    _center_window(modal, parent_win, width, height)
    body = tk.Frame(modal, bg="white", padx=10, pady=10); body.pack(fill="both", expand=True)
    tk.Label(body, text=f"Numéro facture: {invoice_identifier}", font=("Segoe UI",10,"bold"), bg="white").pack(anchor="w", pady=(0,6))
    tk.Label(body, text="Motif d'annulation (obligatoire) :", font=FONT_LABEL, bg="white").pack(anchor="w", pady=(6,2))
    motif_text = tk.Text(body, height=6, wrap="word", bd=1, relief="solid"); motif_text.pack(fill="both", pady=(0,6))

    def _send_cancel():
        motif = motif_text.get("1.0", "end").strip()
        if not motif:
            messagebox.showwarning("Motif requis", "Veuillez saisir un motif d'annulation.", parent=modal); return
        ok, msg = cancel_invoice_api(invoice_identifier, motif)
        if ok:
            try:
                conn = get_connection(); cur = conn.cursor()
                cur.execute("PRAGMA table_info(facture)")
                cols = [c[1] for c in cur.fetchall()]
                if "cn_motif" not in cols:
                    cur.execute("ALTER TABLE facture ADD COLUMN cn_motif TEXT")
                cur.execute("UPDATE facture SET cn_motif = ?, facture_statut = ? WHERE id = ?", (motif, "annulé", facture_id_local))
                conn.commit(); conn.close()
            except Exception:
                logger.exception("Impossible d'enregistrer motif local")
            messagebox.showinfo("Annulation", f"Succès: {msg}", parent=modal)
            modal.destroy()
            try:
                if parent_win:
                    pass
            except Exception:
                pass
        else:
            messagebox.showerror("Erreur annulation", f"{msg}", parent=modal)

    row = tk.Frame(body, bg="white"); row.pack(fill="x", pady=(4,0))
    spacer = tk.Frame(row, bg="white"); spacer.pack(side="left", expand=True)
    b_confirm = ttk.Button(row, text="Confirmer annulation", command=_send_cancel, style="Danger.TButton"); b_confirm.pack(side="right", padx=6)
    b_cancel = ttk.Button(row, text="Annuler", command=modal.destroy, style="Default.TButton"); b_cancel.pack(side="right", padx=6)

# -----------------------
# cancelInvoice API wrapper
# -----------------------
def cancel_invoice_api(invoice_identifier: str, motif: str = None) -> Tuple[bool, str]:
    token = obtenir_token_auto()
    if not token:
        return False, "Token OBR introuvable"
    url = "https://ebms.obr.gov.bi:9443/ebms_api/cancelInvoice/"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    payload = {"invoice_identifier": invoice_identifier}
    if motif:
        payload["cn_motif"] = motif
    try:
        resp = requests.post(url, json=payload, headers=headers, verify=False, timeout=30)
    except Exception as ex:
        logger.exception("Network error cancelInvoice: %s", ex)
        return False, str(ex)
    try:
        j = resp.json()
    except Exception:
        return False, resp.text if resp is not None else "Erreur inconnue"
    success = bool(resp.status_code == 200 and j.get("success"))
    msg = j.get("msg") or j.get("message") or resp.text
    return success, str(msg)


# -----------------------
# Export wrappers (Excel/PDF)
# -----------------------
def export_invoices_excel(rows: List[Dict], items_by_invoice: Dict[int, List[Dict]], filename: str) -> Tuple[bool, Optional[str]]:
    """
    Exporte deux feuilles Excel :
      - Factures : une ligne par facture + colonnes agrégées d'articles (concaténation par '; ')
      - Articles : liste aplatie des articles (une ligne par article)
    Requis : pandas + openpyxl (pd variable) et imports openpyxl.styles, openpyxl.utils.get_column_letter
    """
    if pd is None:
        return False, "pandas/openpyxl non installé"
    try:
        # 1) Enrichir chaque invoice avec les colonnes articles concaténées
        enriched_rows = []
        for inv in rows:
            fid = inv.get("id")
            items = items_by_invoice.get(fid, []) or []

            def join_field(key):
                vals = []
                for it in items:
                    v = it.get(key)
                    if v is None:
                        continue
                    vals.append(str(v))
                return "; ".join(vals) if vals else ""

            row_copy = dict(inv)  # clone pour ne pas muter l'original
            row_copy["item_designation"] = join_field("item_designation")
            row_copy["quantity"] = join_field("quantity")
            row_copy["unit_price_used"] = join_field("unit_price_used")
            row_copy["unit_price_nvat"] = join_field("unit_price_nvat")
            row_copy["vat_amount"] = join_field("vat_amount")
            row_copy["line_total_amount"] = join_field("line_total_amount")
            enriched_rows.append(row_copy)

        # 2) Construire DataFrame Factures avec colonnes désirées (existantes + articles)
        df = pd.DataFrame(enriched_rows)
        if df.empty:
            return False, "Aucune donnée"

        # Colonnes de base à garder (si présentes)
        base_cols = [c for c in ["invoice_number", "invoice_date", "invoice_type", "customer_name", "facture_statut"] if c in df.columns]

        # Assurer les colonnes article présentes (création si manquantes)
        article_cols = ["item_designation", "quantity", "unit_price_used", "unit_price_nvat", "vat_amount", "line_total_amount"]
        for ac in article_cols:
            if ac not in df.columns:
                df[ac] = ""

        # Sélectionner colonnes finales dans l'ordre souhaité
        final_cols = []
        final_cols += base_cols
        final_cols += article_cols

        if not base_cols:
            existing_non_article = [c for c in df.columns if c not in article_cols]
            final_cols = existing_non_article + article_cols

        df_factures = df[final_cols].rename(columns={
            "invoice_number": "N° du facture",
            "invoice_date": "Date",
            "invoice_type": "Type",
            "customer_name": "Client",
            "facture_statut": "Statut",
            "item_designation": "Désignation (articles)",
            "quantity": "Qté (articles)",
            "unit_price_used": "PU utilisé (articles)",
            "unit_price_nvat": "PU HT (articles)",
            "vat_amount": "TVA (articles)",
            "line_total_amount": "Total ligne (articles)"
        })

        # 3) Construire DataFrame Articles (une ligne par article)
        articles_records = []
        for inv in rows:
            fid = inv.get("id")
            inv_num = inv.get("invoice_number")
            inv_date = inv.get("invoice_date")
            inv_type = inv.get("invoice_type")
            for it in items_by_invoice.get(fid, []):
                articles_records.append({
                    "N°": inv_num,
                    "Date": inv_date,
                    "Type": inv_type,
                    "Désignation": it.get("item_designation", ""),
                    "Qté": it.get("quantity", ""),
                    "PU utilisé": it.get("unit_price_used", ""),
                    "PU HT": it.get("unit_price_nvat", ""),
                    "TVA": it.get("vat_amount", ""),
                    "Total ligne": it.get("line_total_amount", ""),
                })
        df_items = pd.DataFrame(articles_records)

        # 4) Écriture Excel avec formatage des en-têtes et ajustement colonnes
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Feuille Factures
            df_factures.to_excel(writer, index=False, sheet_name="Factures")
            ws = writer.book["Factures"]
            hdr_font = Font(bold=True, size=11)
            hdr_fill = PatternFill(start_color="D9E6F6", end_color="D9E6F6", fill_type="solid")
            for col_idx, col in enumerate(df_factures.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                max_len = max((len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)), default=len(col))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, len(col) + 2), 60)

            # Feuille Articles
            if not df_items.empty:
                df_items.to_excel(writer, index=False, sheet_name="Articles")
                ws2 = writer.book["Articles"]
                for col_idx, col in enumerate(df_items.columns, start=1):
                    cell = ws2.cell(row=1, column=col_idx)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    max_len = max((len(str(ws2.cell(r, col_idx).value or "")) for r in range(1, ws2.max_row + 1)), default=len(col))
                    ws2.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, len(col) + 2), 60)

        return True, None

    except Exception as e:
        logger.exception("Export Excel failed: %s", e)
        return False, str(e)


def export_invoices_pdf(rows: List[Dict], items_by_invoice: Dict[int, List[Dict]], filename: str) -> Tuple[bool, Optional[str]]:
    """
    Exporte un PDF (landscape A4) listant les factures puis, pour chaque facture, ses articles.
    Amélioration ciblée : lisibilité du premier tableau (Factures) — wrapping, largeur contrôlée, fontsize réduit.
    """
    if SimpleDocTemplate is None:
        return False, "reportlab non installé"
    try:
        # Page landscape A4, marges contrôlées
        page_w, _ = landscape(A4)
        usable_w = page_w - 36  # marges gauche/droite
        styles = getSampleStyleSheet()

        # Styles dédiés pour le tableau (petit mais lisible)
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8.5, leading=10)
        small_bold = ParagraphStyle("small_bold", parent=styles["Normal"], fontSize=9, leading=10, spaceAfter=2)
        heading = ParagraphStyle("heading", parent=styles["Heading1"], fontSize=14, leading=16)

        doc = SimpleDocTemplate(filename, pagesize=landscape(A4),
                                leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        elems = [Paragraph("Liste des factures (avec détails)", heading), Spacer(1, 8)]

        # Tableau entête Factures — colonnes larges pour Client, compact pour les autres
        header_aliases = ["invoice_number", "invoice_date", "invoice_type", "customer_name", "facture_statut"]
        labels = ["N° du facture", "Date", "Type", "Client", "Statut"]

        # Largeurs équilibrées : favoriser la colonne Client et laisser de la place au statut si nécessaire
        col_widths = [usable_w * 0.12, usable_w * 0.14, usable_w * 0.12, usable_w * 0.46, usable_w * 0.16]

        # Construire table_data en utilisant Paragraphs pour permettre le wrap
        table_data = [[Paragraph(l, small_bold) for l in labels]]
        for r in rows:
            row_vals = []
            for alias in header_aliases:
                v = r.get(alias, "") or ""
                if alias == "invoice_date":
                    v = format_date_short_pdf(v)
                if isinstance(v, str) and len(v) > 220:
                    v = v[:210] + "…"
                para = Paragraph(str(v), small)
                row_vals.append(para)
            table_data.append(row_vals)

        # Table avec style amélioré
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.28, colors.HexColor("#c7d2e7")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E6F6")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ALIGN", (0, 0), (1, -1), "LEFT"),
            ("ALIGN", (1, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (3, -1), "LEFT"),
            ("ALIGN", (4, 0), (4, -1), "CENTER"),
        ])
        tbl.setStyle(tbl_style)

        elems.append(tbl)
        elems.append(Spacer(1, 12))

        # Pour chaque facture, un sous-tableau des articles
        for inv in rows:
            inv_title = f"Facture {inv.get('invoice_number','')} - {format_date_short_pdf(inv.get('invoice_date',''))} - {inv.get('invoice_type','')}"
            elems.append(Paragraph(inv_title, styles["Heading3"]))
            elems.append(Spacer(1, 4))

            items = items_by_invoice.get(inv.get("id"), []) or []

            item_labels = ["Désignation", "Qté", "PU utilisé", "PU HT", "TVA", "Total ligne"]
            item_widths = [usable_w * 0.42, usable_w * 0.10, usable_w * 0.12, usable_w * 0.12, usable_w * 0.12, usable_w * 0.12]

            item_rows = [[Paragraph(l, small_bold) for l in item_labels]]
            for it in items:
                des_para = Paragraph(str(it.get("item_designation", "") or ""), small)
                qty = Paragraph(str(it.get("quantity", "") or ""), small)
                pu_used = Paragraph(str(it.get("unit_price_used", "") or ""), small)
                pu_nvat = Paragraph(str(it.get("unit_price_nvat", "") or ""), small)
                vat_amt = Paragraph(str(it.get("vat_amount", "") or ""), small)
                line_tot = Paragraph(str(it.get("line_total_amount", "") or ""), small)
                item_rows.append([des_para, qty, pu_used, pu_nvat, vat_amt, line_tot])

            item_tbl = Table(item_rows, colWidths=item_widths, repeatRows=1)
            item_tbl.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.24, colors.HexColor("#d1d5db")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF4FF")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ]))
            elems.append(item_tbl)
            elems.append(Spacer(1, 12))

        # Build PDF
        doc.build(elems)
        return True, None

    except Exception as e:
        logger.exception("Export PDF failed: %s", e)
        return False, str(e)


# -----------------------
# afficher_liste_factures complet (UI)
# -----------------------
def afficher_liste_factures(parent, page_size: Optional[int] = None):
    """
    Affiche la liste des factures dans l'interface.
    - parent : widget Tk parent
    - page_size : optionnel, remplace la taille de page par défaut (PAGE_SIZE)
    """
    
    from datetime import date, timedelta

    # try to access global session (role)
    try:
        from models.session import session as global_session
    except Exception:
        global_session = None

    def _current_user_can_manage():
        try:
            if not global_session:
                return False
            role = getattr(global_session, "role", None)
            if role is None:
                return False
            return str(role).lower() in ("admin", "manager")
        except Exception:
            return False

    # clear parent
    for w in parent.winfo_children():
        try:
            w.destroy()
        except Exception:
            pass
    try:
        parent.configure(bg=CONTENT_BG)
    except Exception:
        pass

    # determine effective page size (param override -> env/config -> default)
    effective_page_size = PAGE_SIZE if page_size in (None, 0) else int(page_size)
    effective_page_size = max(1, min(effective_page_size, 500))

    # default dates: yesterday -> tomorrow in the inputs (per request)
    today = date.today()
    yesterday = today - timedelta(days=1)
    tomorrow = today + timedelta(days=1)
    yesterday_str = yesterday.isoformat()
    tomorrow_str = tomorrow.isoformat()

    state = {
        "data": [],
        "page": 1,
        "page_size": effective_page_size,
        "total_pages": 1,
        "date_from": None,
        "date_to": None,
        "status_filter": "ALL"
    }

    # Header / Controls
    title_frame = tk.Frame(parent, bg=CONTENT_BG)
    title_frame.pack(fill="x", padx=12, pady=(12, 6))
    tk.Label(title_frame, text="📄 Liste des factures", font=FONT_TITLE, bg=CONTENT_BG, fg=TITLE_FG).pack(anchor="w")

    ctrl = tk.Frame(parent, bg=CONTENT_BG)
    ctrl.pack(fill="x", padx=12, pady=(0, 8))

    tk.Label(ctrl, text="Date de :", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).grid(row=0, column=0, sticky="w", padx=(0, 6))
    date_from_var = tk.StringVar(value=yesterday_str)
    if DateEntry:
        date_entry_from = DateEntry(ctrl, textvariable=date_from_var, date_pattern='yyyy-mm-dd', width=12)
    else:
        date_entry_from = tk.Entry(ctrl, textvariable=date_from_var, width=12)
    date_entry_from.grid(row=0, column=1, padx=(0, 12))

    tk.Label(ctrl, text="Date à :", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).grid(row=0, column=2, sticky="w", padx=(0, 6))
    date_to_var = tk.StringVar(value=tomorrow_str)
    if DateEntry:
        date_entry_to = DateEntry(ctrl, textvariable=date_to_var, date_pattern='yyyy-mm-dd', width=12)
    else:
        date_entry_to = tk.Entry(ctrl, textvariable=date_to_var, width=12)
    date_entry_to.grid(row=0, column=3, padx=(0, 12))

    # "Aujourd'hui" quick-reset button (styled)
    def _reset_to_today_range():
        # set fields to yesterday..tomorrow as requested
        date_from_var.set(yesterday_str)
        date_to_var.set(tomorrow_str)
        load_data(page=1)

    btn_today = ttk.Button(ctrl, text="Aujourd'hui", command=_reset_to_today_range, style="Primary.TButton")
    # add a distinctive background using the configured style; keep visual cue
    btn_today.grid(row=0, column=4, padx=(0, 10))

    tk.Label(ctrl, text="Statut :", font=FONT_LABEL, bg=CONTENT_BG, fg=LABEL_FG).grid(row=0, column=5, sticky="w", padx=(0, 6))
    status_var = tk.StringVar(value="ALL")
    cb = ttk.Combobox(ctrl, textvariable=status_var, values=["ALL", "envoyé", "non_envoyé", "annulé"], width=14, state="readonly")
    cb.grid(row=0, column=6, padx=(0, 12))

    row2 = tk.Frame(ctrl, bg=CONTENT_BG)
    row2.grid(row=1, column=0, columnspan=7, sticky="w", pady=(8, 0))
    btn_search = ttk.Button(row2, text="Recherche", style="Primary.TButton")
    btn_search.grid(row=0, column=0, padx=(0, 6))
    btn_refresh = ttk.Button(row2, text="Rafraîchir", style="Default.TButton")
    btn_refresh.grid(row=0, column=1, padx=(0, 6))
    btn_export_xls = ttk.Button(row2, text="Export Excel", style="Success.TButton")
    btn_export_xls.grid(row=0, column=2, padx=(0, 6))
    btn_export_pdf = ttk.Button(row2, text="Export PDF", style="Success.TButton")
    btn_export_pdf.grid(row=0, column=3, padx=(0, 6))

    # disable export buttons for non-admin/manager (UI-level)
    if not _current_user_can_manage():
        try:
            btn_export_xls.state(["disabled"])
            btn_export_pdf.state(["disabled"])
        except Exception:
            try:
                btn_export_xls.config(state="disabled")
                btn_export_pdf.config(state="disabled")
            except Exception:
                pass

    # Card / Table area
    card = tk.Frame(parent, bg=CARD_BG)
    card.pack(fill="both", padx=12, pady=(0, 8), expand=True)
    inner_outer = tk.Frame(card, bg=CONTOUR_BG, bd=2, relief="groove")
    inner_outer.pack(fill="both", expand=True, padx=8, pady=8)
    body_inner = tk.Frame(inner_outer, bg=CARD_BG)
    body_inner.pack(fill="both", expand=True, padx=4, pady=4)

    total_cols = len(INVOICE_COLUMNS)
    CLIENT_COL_IDX = next((i for i, c in enumerate(INVOICE_COLUMNS) if c[0] == "customer_name"), None)
    for c, col in enumerate(INVOICE_COLUMNS):
        alias, label, w = col
        hdr = tk.Label(body_inner, text=label, bg=TABLE_HEADER_BG, fg=LABEL_FG, font=TABLE_HEADER_FONT, anchor="w", bd=1, relief="solid", padx=TABLE_CELL_PADX, pady=TABLE_CELL_PADY)
        hdr.grid(row=0, column=c, sticky="nsew")
        if alias == "customer_name":
            body_inner.grid_columnconfigure(c, weight=3, minsize=140)
        else:
            body_inner.grid_columnconfigure(c, weight=1, minsize=90)
    actions_hdr = tk.Label(body_inner, text="Actions", bg=TABLE_HEADER_BG, fg=LABEL_FG, font=TABLE_HEADER_FONT, bd=1, relief="solid", padx=TABLE_CELL_PADX, pady=TABLE_CELL_PADY)
    actions_hdr.grid(row=0, column=total_cols, sticky="nsew")
    body_inner.grid_columnconfigure(total_cols, weight=0, minsize=200)

    # Pager
    pager = tk.Frame(parent, bg=CONTENT_BG)
    pager.pack(fill="x", padx=12, pady=(0, 12))
    lbl_page_info = tk.Label(pager, text="", bg=CONTENT_BG, fg=LABEL_FG, font=FONT_LABEL)
    lbl_page_info.pack(side="right", padx=(8, 12))
    btn_prev = ttk.Button(pager, text="Précédent", state="disabled")
    btn_prev.pack(side="left", padx=(0, 8))
    btn_next = ttk.Button(pager, text="Suivant", state="disabled")
    btn_next.pack(side="left", padx=(0, 8))

    row_widgets: List[List[tk.Widget]] = []

    # Data load
    def load_data(page: int = 1):
        # if user erased fields, fallback to yesterday..tomorrow
        raw_from = date_from_var.get().strip() or yesterday_str
        raw_to = date_to_var.get().strip() or tomorrow_str
        state["date_from"] = parse_date_input(raw_from)
        state["date_to"] = parse_date_input(raw_to)
        state["status_filter"] = status_var.get() if status_var.get() else "ALL"
        try:
            conn = get_connection()
            cur = conn.cursor()
            q = """
                SELECT f.id, f.invoice_number, f.invoice_date, f.invoice_type, f.invoice_identifier,
                       COALESCE(c.customer_name, '') as customer_name,
                       f.facture_statut
                FROM facture f
                LEFT JOIN client c ON f.client_id = c.id
                WHERE 1=1
            """
            params = []
            if state["date_from"]:
                q += " AND f.invoice_date >= ?"
                params.append(state["date_from"])
            if state["date_to"]:
                q += " AND f.invoice_date <= ?"
                params.append(state["date_to"])
            if state["status_filter"] and state["status_filter"] != "ALL":
                q += " AND f.facture_statut = ?"
                params.append(state["status_filter"])
            q += " ORDER BY f.id DESC"
            cur.execute(q, params)
            rows = cur.fetchall() or []
            conn.close()
            cols = ["id", "invoice_number", "invoice_date", "invoice_type", "invoice_identifier", "customer_name", "facture_statut"]
            state["data"] = []
            for r in rows:
                rt = tuple(r)
                d = dict(zip(cols, rt if len(rt) >= len(cols) else tuple(list(rt) + [None] * (len(cols) - len(rt)))))
                state["data"].append(d)
        except Exception as e:
            logger.exception("Load invoices failed: %s", e)
            state["data"] = []

        total = len(state["data"])
        state["page_size"] = effective_page_size
        state["total_pages"] = max(1, math.ceil(total / state["page_size"]))
        state["page"] = max(1, min(page, state["total_pages"]))
        refresh_table()

    # Refresh UI table
    def refresh_table():
        for wr in row_widgets:
            for w in wr:
                try:
                    w.destroy()
                except Exception:
                    pass
        row_widgets.clear()
        for w in body_inner.winfo_children():
            try:
                info = w.grid_info()
                if int(info.get("row", 0)) > 0:
                    w.destroy()
            except Exception:
                pass

        if not state["data"]:
            lbl_empty = tk.Label(body_inner, text="Aucune facture pour les filtres choisis.", font=("Segoe UI", 12), bg=CARD_BG, fg=LABEL_FG, padx=20, pady=20)
            lbl_empty.grid(row=1, column=0, columnspan=len(INVOICE_COLUMNS) + 1, sticky="nsew")
            lbl_page_info.config(text="Page 0 / 0")
            btn_prev.config(state="disabled")
            btn_next.config(state="disabled")
            return

        page = state["page"]
        size = state["page_size"]
        start = (page - 1) * size
        end = start + size
        page_rows = state["data"][start:end]

        for ridx, row in enumerate(page_rows, start=1):
            widgets_row: List[tk.Widget] = []
            bg = ROW_ALT if ridx % 2 == 0 else CARD_BG
            for cidx, col in enumerate(INVOICE_COLUMNS):
                key = col[0]
                val = row.get(key, "")
                if key == "invoice_date":
                    val = format_date_short(val)
                lbl_text = str(val or "")
                wraplen = 320 if cidx == CLIENT_COL_IDX else 200
                cell = tk.Label(body_inner, text=lbl_text, bg=bg, fg=LABEL_FG, font=TABLE_CELL_FONT, anchor="w",
                                bd=1, relief="solid", padx=TABLE_CELL_PADX, pady=TABLE_CELL_PADY, wraplength=wraplen, justify="left")
                cell.grid(row=ridx, column=cidx, sticky="nsew")
                widgets_row.append(cell)

            act_frame = tk.Frame(body_inner, bg=bg, bd=1, relief="solid", padx=4, pady=4)
            act_frame.grid(row=ridx, column=len(INVOICE_COLUMNS), sticky="nsew")

            def _view(r=row):
                _view_invoice(r, parent)

            def _retry(r=row):
                rid = r.get("id")
                if rid:
                    retry_invoice_local_and_send(rid, parent)
                else:
                    messagebox.showerror("Erreur", "ID facture manquant.", parent=parent)

            btn_view = ttk.Button(act_frame, text="Voir", command=_view, style="Primary.TButton")
            btn_view.pack(side="left", padx=4)

            statut = str(row.get("facture_statut", "")).strip().lower()
            invoice_identifier = row.get("invoice_identifier") or ""

            # GetInvoice visible only when facture is 'envoyé'
            if statut in ("envoyé", "envoye", "envoyé"):
                def _getinv(iid=invoice_identifier, p=parent):
                    if not iid:
                        messagebox.showinfo("GetInvoice", "Identifiant OBR manquant pour cette facture.", parent=p)
                        return
                    get_invoice_details(iid, parent=p)
                btn_get = ttk.Button(act_frame, text="GetInvoice", command=_getinv, style="Default.TButton")
                btn_get.pack(side="left", padx=4)

            # Annuler if sent, otherwise Réessayer
            if statut in ("envoyé", "envoye", "envoyé"):
                if invoice_identifier:
                    def _cancel_guard(iid=invoice_identifier, fid=row.get("id"), p=parent):
                        if not _current_user_can_manage():
                            messagebox.showerror("Permission", "Vous n'êtes pas autorisé à annuler des factures.", parent=p)
                            return
                        _open_cancel_modal_with_send(iid, fid, p)
                    btn_cancel = ttk.Button(act_frame, text="Annuler", command=_cancel_guard, style="Danger.TButton")
                    if not _current_user_can_manage():
                        try:
                            btn_cancel.state(["disabled"])
                        except Exception:
                            btn_cancel.config(state="disabled")
                    btn_cancel.pack(side="left", padx=4)
                else:
                    btn_cancel = ttk.Button(act_frame, text="Annuler", command=lambda: messagebox.showinfo("Info", "Identifiant OBR manquant pour l'annulation."), style="Danger.TButton")
                    btn_cancel.pack(side="left", padx=4)
            else:
                btn_retry = ttk.Button(act_frame, text="Réessayer", command=_retry, style="Warning.TButton")
                btn_retry.pack(side="left", padx=4)

            widgets_row.append(act_frame)
            row_widgets.append(widgets_row)

        lbl_page_info.config(text=f"Page {state['page']} / {state['total_pages']}")
        btn_prev.config(state="normal" if state["page"] > 1 else "disabled")
        btn_next.config(state="normal" if state["page"] < state["total_pages"] else "disabled")

    # exporter helper
    def do_export(format: str, date_from_raw: str, date_to_raw: str, parent_win, status_filter_val: str):
        # permission check (defensive)
        if not _current_user_can_manage():
            messagebox.showerror("Permission", "Vous n'êtes pas autorisé à exporter les factures.", parent=parent_win)
            return
        df = parse_date_input(date_from_raw) or parse_date_input(yesterday_str)
        dt = parse_date_input(date_to_raw) or parse_date_input(tomorrow_str)
        status = status_filter_val if status_filter_val else "ALL"
        try:
            conn = get_connection()
            cur = conn.cursor()
            q = """
                SELECT f.id, f.invoice_number, f.invoice_date, f.invoice_type, f.invoice_identifier,
                       COALESCE(c.customer_name, '') as customer_name,
                       f.facture_statut
                FROM facture f
                LEFT JOIN client c ON f.client_id = c.id
                WHERE 1=1
            """
            params = []
            if df:
                q += " AND f.invoice_date >= ?"
                params.append(df)
            if dt:
                q += " AND f.invoice_date <= ?"
                params.append(dt)
            if status and status != "ALL":
                q += " AND f.facture_statut = ?"
                params.append(status)
            q += " ORDER BY f.id DESC"
            cur.execute(q, params)
            rows = cur.fetchall() or []
            cols = ["id", "invoice_number", "invoice_date", "invoice_type", "invoice_identifier", "customer_name", "facture_statut"]
            rows_dict = [dict(zip(cols, tuple(r) if isinstance(r, (list, tuple)) else r)) for r in rows]

            # collect items grouped by facture id
            items_by_invoice = {}
            if rows_dict:
                ids = [r["id"] for r in rows_dict if r.get("id") is not None]
                if ids:
                    cur.execute("PRAGMA table_info(article)")
                    cols_info = cur.fetchall() or []
                    existing_cols = {c[1] for c in cols_info}
                    wanted = ["item_designation", "quantity", "unit_price_used", "unit_price_nvat", "vat_amount", "line_total_amount", "item_code"]
                    select_cols = [c for c in wanted if c in existing_cols]
                    if select_cols:
                        sel = ", ".join(["facture_id"] + select_cols)
                    else:
                        sel = "facture_id, id"
                    cur.execute(f"SELECT {sel} FROM article WHERE facture_id IN ({','.join('?' for _ in ids)}) ORDER BY facture_id, id ASC", tuple(ids))
                    arts = cur.fetchall() or []
                    for a in arts:
                        a = tuple(a)
                        facture_id = a[0]
                        row_map = {}
                        for i, coln in enumerate(select_cols, start=1):
                            row_map[coln] = a[i] if i < len(a) else None
                        items_by_invoice.setdefault(facture_id, []).append(row_map)
            conn.close()

            if format.lower() in ("xlsx", "xls", "excel"):
                path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Liste_Factures.xlsx", parent=parent_win)
                if not path:
                    return
                ok, err = export_invoices_excel(rows_dict, items_by_invoice, path)
                if ok:
                    messagebox.showinfo("Export Excel", f"Export réussi: {os.path.basename(path)}", parent=parent_win)
                else:
                    messagebox.showerror("Export Excel", f"Échec export: {err}", parent=parent_win)
            elif format.lower() == "pdf":
                path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile="Liste_Factures.pdf", parent=parent_win)
                if not path:
                    return
                ok, err = export_invoices_pdf(rows_dict, items_by_invoice, path)
                if ok:
                    messagebox.showinfo("Export PDF", f"Export réussi: {os.path.basename(path)}", parent=parent_win)
                else:
                    messagebox.showerror("Export PDF", f"Échec export: {err}", parent=parent_win)
            else:
                messagebox.showwarning("Export", "Format d'export inconnu.", parent=parent_win)
        except Exception as e:
            logger.exception("Export failed: %s", e)
            messagebox.showerror("Export", f"Erreur lors de l'export: {e}", parent=parent_win)

    # Bind buttons
    btn_search.config(command=lambda: load_data(page=1))
    btn_refresh.config(command=lambda: load_data(page=state.get("page", 1)))
    btn_prev.config(command=lambda: (state.update(page=max(1, state["page"] - 1)), refresh_table()))
    btn_next.config(command=lambda: (state.update(page=min(state["total_pages"], state["page"] + 1)), refresh_table()))
    btn_export_xls.config(command=lambda: do_export("xlsx", date_from_var.get(), date_to_var.get(), parent, status_var.get()))
    btn_export_pdf.config(command=lambda: do_export("pdf", date_from_var.get(), date_to_var.get(), parent, status_var.get()))

    # initial load (defaults to yesterday..tomorrow range)
    load_data(page=1)


# -----------------------
# Démonstration si exécuté directement
# -----------------------
if __name__ == "__main__":
    root = tk.Tk(); root.title("Liste des factures - Demo"); root.geometry("1200x760")
    afficher_liste_factures(root, page_size=20)
    root.mainloop()
  